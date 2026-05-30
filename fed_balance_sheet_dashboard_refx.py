
#### 
####
### 


"""
美联储资产负债表交互式Dashboard
国信证券宏观研究
"""
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl import load_workbook
import numpy as np
from datetime import datetime
import base64
import requests
from io import BytesIO

# 页面配置
st.set_page_config(
    page_title="美联储资产负债表",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 自定义CSS样式 - 参考industry_analysis模板配色
st.markdown("""
<style>
    /* 导入黑体字体 */
    @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+SC:wght@400;500;700&display=swap');

    /* 全局样式 */
    .main {
        background-color: #f0f2f5;
    }

    /* 标题样式 */
    .main-header {
        font-family: 'Noto Sans SC', 'SimHei', sans-serif;
        font-size: 2.2rem;
        font-weight: 700;
        color: #1a1f36;
        text-align: center;
        margin-bottom: 0.5rem;
        padding-top: 1rem;
    }

    .sub-header {
        font-family: 'Noto Sans SC', 'SimHei', sans-serif;
        font-size: 0.9rem;
        color: #666;
        text-align: center;
        margin-bottom: 1.5rem;
    }

    /* 侧边栏样式 - 深蓝色渐变 */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1a1f36 0%, #2d3555 100%) !important;
    }

    section[data-testid="stSidebar"] .stMarkdown,
    section[data-testid="stSidebar"] label,
    section[data-testid="stSidebar"] .stSelectbox label,
    section[data-testid="stSidebar"] .stRadio label {
        color: rgba(255,255,255,0.7) !important;
    }

    section[data-testid="stSidebar"] h1,
    section[data-testid="stSidebar"] h2,
    section[data-testid="stSidebar"] h3 {
        color: white !important;
    }

    /* 侧边栏标题 */
    .sidebar-title {
        color: white;
        font-size: 16px;
        font-weight: 600;
        padding: 0 20px 24px;
        border-bottom: 1px solid rgba(255,255,255,0.1);
        margin-bottom: 20px;
    }

    .sidebar-subtitle {
        color: rgba(255,255,255,0.5);
        font-size: 11px;
        margin-top: 4px;
    }

    /* 卡片样式 */
    .metric-card {
        background: white;
        border-radius: 12px;
        padding: 20px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
        margin-bottom: 16px;
    }

    /* 指标卡片 */
    div[data-testid="stMetric"] {
        background: white;
        border-radius: 12px;
        padding: 16px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }

    div[data-testid="stMetric"] label {
        font-size: 11px;
        color: #888;
    }

    div[data-testid="stMetric"] [data-testid="stMetricValue"] {
        font-size: 1.5rem;
        font-weight: 700;
        color: rgb(4, 68, 119);
    }

    div[data-testid="stMetric"] [data-testid="stMetricDelta"] {
        font-size: 0.9rem;
    }

    /* Tab样式 */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        background: transparent;
    }

    .stTabs [data-baseweb="tab"] {
        padding: 12px 24px;
        background: white;
        border-radius: 8px 8px 0 0;
        font-weight: 500;
        color: #666;
        border: 1px solid #e8e8e8;
        border-bottom: none;
    }

    .stTabs [aria-selected="true"] {
        background: white;
        color: rgb(4, 68, 119);
        font-weight: 600;
        border-bottom: 2px solid rgb(4, 68, 119);
    }

    .stTabs [data-baseweb="tab-panel"] {
        background: white;
        border-radius: 0 0 12px 12px;
        padding: 24px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }

    /* 控制面板 */
    .control-panel {
        background: white;
        border-radius: 12px;
        padding: 20px;
        margin-bottom: 20px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }

    /* 图表标题 */
    .chart-title {
        font-size: 15px;
        font-weight: 600;
        color: #1a1f36;
        margin-bottom: 16px;
    }

    /* 按钮样式 */
    .stButton button {
        background: rgb(4, 68, 119);
        color: white;
        border-radius: 8px;
        font-weight: 500;
        border: none;
        padding: 10px 20px;
    }

    .stButton button:hover {
        background: rgb(6, 85, 148);
        transform: translateY(-1px);
        box-shadow: 0 4px 12px rgba(4, 68, 119, 0.25);
    }

    /* 轴选择按钮样式 - 小巧透明 */
    .stButton button[kind="secondary"] {
        background: rgba(4, 68, 119, 0.15) !important;
        color: rgb(4, 68, 119) !important;
        border-radius: 4px !important;
        font-weight: 400 !important;
        padding: 2px 8px !important;
        font-size: 11px !important;
        border: 1px solid rgba(4, 68, 119, 0.2) !important;
        min-height: 22px !important;
        line-height: 1 !important;
    }

    .stButton button[kind="secondary"]:hover {
        background: rgba(4, 68, 119, 0.25) !important;
        transform: none !important;
        box-shadow: none !important;
    }

    /* 下拉选择框 */
    .stMultiSelect, .stSelectbox {
        background: white;
    }

    /* 页脚样式 */
    .footer {
        text-align: center;
        padding: 20px;
        color: #888;
        font-size: 12px;
        margin-top: 20px;
        border-top: 1px solid #e8e8e8;
    }

    /* Logo容器 */
    .logo-container {
        text-align: center;
        padding: 20px;
        background: white;
        border-radius: 12px;
        margin-top: 20px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }

    /* 侧边栏logo */
    .sidebar-logo {
        text-align: center;
        padding: 20px 0;
        border-bottom: 1px solid rgba(255,255,255,0.1);
        margin-bottom: 20px;
    }

    /* 隐藏streamlit默认元素 */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}

    /* 去掉expander边框 */
    .stExpander {
        border: none !important;
        box-shadow: none !important;
    }
    .stExpander > details {
        border: none !important;
        background: transparent !important;
    }
    .stExpander > details > summary {
        border: none !important;
        background: transparent !important;
        padding: 4px 0 !important;
    }
    .stExpander > details > div {
        border: none !important;
        background: transparent !important;
        padding: 0 !important;
    }
</style>
""", unsafe_allow_html=True)


def img_to_base64(path_or_url):
    """将图片转换为base64编码，支持本地路径或URL"""
    if path_or_url.startswith('http'):
        # 从URL加载
        response = requests.get(path_or_url, timeout=10)
        response.raise_for_status()
        data = response.content
    else:
        # 从本地文件加载
        with open(path_or_url, "rb") as f:
            data = f.read()
    return base64.b64encode(data).decode()


@st.cache_data
def load_data():
    """加载Excel数据，支持本地路径或URL"""
    # 数据文件URL（可替换为本地路径）
    data_url = 'https://mengke25.github.io/guosen/global_market/convert/【国信宏观】美联储资产负债（自动更新）.xlsx'

    # 从URL下载Excel文件
    response = requests.get(data_url, timeout=30)
    response.raise_for_status()

    # 使用BytesIO作为文件对象
    wb = load_workbook(BytesIO(response.content), data_only=True)

    # 读取资产数据
    ws_asset = wb['资产']
    asset_data = []
    for row in ws_asset.iter_rows(min_row=9, values_only=True):
        if row[0] is not None:
            asset_data.append(row)

    # 获取资产列名
    asset_headers = [cell.value for cell in ws_asset[7] if cell.value is not None]

    df_asset = pd.DataFrame(asset_data)
    df_asset = df_asset.iloc[:, :len(asset_headers)]
    df_asset.columns = asset_headers

    # 简化列名
    asset_col_map = {
        '日期': '日期',
        '美国:所有联储银行:资产:总资产': '总资产',
        '美国:所有联储银行:资产:黄金证券账户': '黄金证券账户',
        '美国:所有联储银行:资产:特别提款权账户': '特别提款权账户',
        '美国:所有联储银行:资产:硬币': '硬币',
        '美国:所有联储银行:资产:持有证券、未摊销证券溢价和折扣、回购协议和贷款': '持有证券及贷款总额',
        '美国:所有联储银行:资产:持有证券': '持有证券',
        '美国:所有联储银行:资产:持有证券:美国国债': '美国国债',
        '美国:所有联储银行:资产:持有证券:美国国债:短期债券': '国债-短期债券',
        '美国:所有联储银行:资产:持有证券:美国国债:中长期名义债券': '国债-中长期名义债券',
        '美国:所有联储银行:资产:持有证券:美国国债:中长期通胀指数债券': '国债-中长期通胀指数债券',
        '美国:所有联储银行:资产:持有证券:美国国债:通胀补偿': '国债-通胀补偿',
        '美国:所有联储银行:资产:持有证券:美国国债:15天以内': '国债-15天以内',
        '美国:所有联储银行:资产:持有证券:美国国债:16-90天': '国债-16-90天',
        '美国:所有联储银行:资产:持有证券:美国国债:91天-1年': '国债-91天-1年',
        '美国:所有联储银行:资产:持有证券:美国国债:1-5年': '国债-1-5年',
        '美国:所有联储银行:资产:持有证券:美国国债:5-10年': '国债-5-10年',
        '美国:所有联储银行:资产:持有证券:美国国债:大于10年': '国债-大于10年',
        '美国:所有联储银行:资产:持有证券:联邦机构债券': '联邦机构债券',
        '美国:所有联储银行:资产:持有证券:抵押贷款支持债券(MBS)': 'MBS',
        '美国:所有联储银行:资产:当前持有的未摊销证券溢价': '未摊销证券溢价',
        '美国:所有联储银行:资产:当前持有的未摊销证券折扣': '未摊销证券折扣',
        '美国:所有联储银行:资产:正向回购协议': '正向回购协议',
        '美国:所有联储银行:资产:其他贷款': '其他贷款',
        '美国:所有联储银行:资产:储备银行信贷:MS Facilities LLC的净投资组合持有量(中产阶级贷款计划)': 'MSF投资组合',
        '美国:所有联储银行:资产:储备银行信贷:市政流动性基金有限责任公司持有的净投资组合': '市政流动性基金',
        '美国:所有联储银行:资产:持有TALF II LLC投资组合净额': 'TALF II投资组合',
        '美国:所有联储银行:资产:托收中项目': '托收中项目',
        '美国:所有联储银行:资产:银行不动产': '银行不动产',
        '美国:所有联储银行:资产:中央银行流动性互换': '中央银行流动性互换',
        '美国:所有联储银行:资产:外币': '外币',
        '美国:所有联储银行:资产:其他联储资产': '其他资产'
    }
    df_asset = df_asset.rename(columns=lambda x: asset_col_map.get(x, x))

    # 读取负债数据
    ws_liab = wb['负债']
    liab_data = []
    for row in ws_liab.iter_rows(min_row=6, values_only=True):
        if row[0] is not None:
            liab_data.append(row)

    # 获取负债列名
    liab_headers = [cell.value for cell in ws_liab[4] if cell.value is not None]

    df_liab = pd.DataFrame(liab_data)
    df_liab = df_liab.iloc[:, :len(liab_headers)]
    df_liab.columns = liab_headers

    # 简化列名
    liab_col_map = {
        '日期': '日期',
        '美国:所有联储银行:负债:总负债': '总负债',
        '美国:所有联储银行:负债:存款': '存款',
        '美国:所有联储银行:负债:存款:存款机构定期存款': '存款-定期存款',
        '美国:所有联储银行:负债:存款:存款机构其他存款': '存款-其他存款',
        '美国:所有联储银行:负债:存款:美国财政部一般账户': '财政部一般账户',
        '美国:所有联储银行:负债:存款:外国官方': '外国官方存款',
        '美国:所有联储银行:负债:存款:其他存款': '其他存款',
        '美国:所有联储银行:负债:逆向回购协议': '逆向回购协议',
        '美国:所有联储银行:负债:联储票据(扣除美联储自持部分)': '联储票据',
        '美国:所有联储银行:负债:延迟入账现金项目': '延迟入账现金',
        '美国:所有联储银行:负债:财政部对信贷部门的捐助': '财政部捐助',
        '美国:所有联储银行:负债:其他负债及应计股息': '其他负债',
        '美国:所有联储银行:负债:资本账户:实缴资本': '实缴资本',
        '美国:所有联储银行:负债:资本账户:结余': '资本结余',
        '美国:所有联储银行:负债:资本账户:其他资本账户': '其他资本',
        '全部资本:净资产': '净资产'
    }
    df_liab = df_liab.rename(columns=lambda x: liab_col_map.get(x, x))

    # 处理日期列
    df_asset['日期'] = pd.to_datetime(df_asset['日期'])
    df_liab['日期'] = pd.to_datetime(df_liab['日期'])

    # 删除重复日期，保留最后一个
    df_asset = df_asset.drop_duplicates(subset='日期', keep='last')
    df_liab = df_liab.drop_duplicates(subset='日期', keep='last')

    # 数据类型转换
    for col in df_asset.columns:
        if col != '日期':
            df_asset[col] = pd.to_numeric(df_asset[col], errors='coerce')

    for col in df_liab.columns:
        if col != '日期':
            df_liab[col] = pd.to_numeric(df_liab[col], errors='coerce')

    # 排序
    df_asset = df_asset.sort_values('日期').reset_index(drop=True)
    df_liab = df_liab.sort_values('日期').reset_index(drop=True)

    return df_asset, df_liab


# ==================== 数据变换函数 ====================
def transform_yoy(dates, values):
    """计算同比变化率 (%)"""
    if len(values) < 53:  # 需要至少53周数据
        return dates, values
    yoy_dates = []
    yoy_values = []
    for i in range(52, len(values)):
        if values[i-52] != 0:
            yoy = ((values[i] - values[i-52]) / abs(values[i-52])) * 100
            yoy_dates.append(dates[i])
            yoy_values.append(yoy)
    return yoy_dates, yoy_values


def transform_mom(dates, values):
    """计算环比变化率 (%)"""
    if len(values) < 5:
        return dates, values
    mom_dates = []
    mom_values = []
    for i in range(4, len(values)):
        if values[i-4] != 0:
            mom = ((values[i] - values[i-4]) / abs(values[i-4])) * 100
            mom_dates.append(dates[i])
            mom_values.append(mom)
    return mom_dates, mom_values


def transform_log(dates, values):
    """计算对数值"""
    import math
    log_dates = []
    log_values = []
    for date, value in zip(dates, values):
        if value > 0:
            log_dates.append(date)
            log_values.append(math.log(value))
    return log_dates, log_values


def apply_transform(dates, values, transform_type):
    """应用数据变换"""
    if transform_type == 'yoy':
        return transform_yoy(dates, values)
    elif transform_type == 'mom':
        return transform_mom(dates, values)
    elif transform_type == 'log':
        return transform_log(dates, values)
    else:
        return dates, values


def resample_data(dates, values, frequency, agg_method='mean'):
    """变频聚合数据

    Args:
        dates: 日期列表
        values: 数值列表
        frequency: 目标频率 - 'M'(月度), 'Q'(季度), 'A'(年度)
        agg_method: 聚合方法 - 'mean', 'sum', 'last'
    """
    if not dates or not values or frequency == 'raw':
        return dates, values

    from collections import defaultdict

    # 按周期分组
    grouped = defaultdict(list)
    for date, value in zip(dates, values):
        if frequency == 'M':  # 月度
            key = (date.year, date.month)
        elif frequency == 'Q':  # 季度
            key = (date.year, (date.month - 1) // 3 + 1)
        elif frequency == 'A':  # 年度
            key = (date.year,)
        else:
            return dates, values
        grouped[key].append((date, value))

    # 聚合
    resampled_dates = []
    resampled_values = []

    for key in sorted(grouped.keys()):
        group = grouped[key]
        group_values = [v for _, v in group]

        # 使用该周期最后一个日期
        rep_date = max(d for d, _ in group)

        if agg_method == 'mean':
            agg_value = sum(group_values) / len(group_values)
        elif agg_method == 'sum':
            agg_value = sum(group_values)
        elif agg_method == 'last':
            agg_value = group_values[-1]
        else:
            agg_value = sum(group_values) / len(group_values)

        resampled_dates.append(rep_date)
        resampled_values.append(agg_value)

    return resampled_dates, resampled_values


def create_chart_layout(title, y_title="金额 (百万美元)", x_title="日期"):
    """创建统一的图表布局样式 - 黑体字体，0.5磅轴线"""
    return {
        'title': {
            'text': title,
            'font': {'family': 'SimHei, Noto Sans SC, sans-serif', 'size': 16, 'color': '#1a1f36'},
            'x': 0.5,
            'xanchor': 'center'
        },
        'xaxis': {
            'title': {'text': x_title, 'font': {'family': 'SimHei, Noto Sans SC, sans-serif', 'size': 12}},
            'tickfont': {'family': 'SimHei, Noto Sans SC, sans-serif', 'size': 11},
            'linecolor': '#333',
            'linewidth': 0.5,
            'mirror': False,
            'ticks': 'inside',
            'ticklen': 4,
            'gridcolor': '#e8e8e8',
            'gridwidth': 0.5
        },
        'yaxis': {
            'title': {'text': y_title, 'font': {'family': 'SimHei, Noto Sans SC, sans-serif', 'size': 12}},
            'tickfont': {'family': 'SimHei, Noto Sans SC, sans-serif', 'size': 11},
            'linecolor': '#333',
            'linewidth': 0.5,
            'mirror': False,
            'ticks': 'inside',
            'ticklen': 4,
            'gridcolor': '#e8e8e8',
            'gridwidth': 0.5
        },
        'hovermode': 'x unified',
        'plot_bgcolor': '#fafafa',
        'paper_bgcolor': 'white',
        'margin': {'l': 60, 'r': 30, 't': 50, 'b': 60},
        'legend': {
            'font': {'family': 'SimHei, Noto Sans SC, sans-serif', 'size': 11},
            'orientation': 'h',
            'yanchor': 'bottom',
            'y': 1.02,
            'xanchor': 'right',
            'x': 1
        }
    }


def format_large_number(num):
    """格式化大数字显示"""
    if pd.isna(num):
        return "-"
    if abs(num) >= 1e6:
        return f"{num/1e6:.2f}万亿"
    elif abs(num) >= 1e4:
        return f"{num/1e4:.2f}万亿"
    else:
        return f"{num:,.0f}"


# 加载图片（使用URL，方便云端部署）
logo_path = 'https://mengke25.github.io/guosen/global_market/convert/logo-bgremover.png'
qrcode_path = 'https://mengke25.github.io/guosen/global_market/convert/qrcode_for_gh_e27313512349_430.jpg'

try:
    logo_base64 = img_to_base64(logo_path)
    qrcode_base64 = img_to_base64(qrcode_path)
    images_loaded = True
except:
    images_loaded = False

# 加载数据
try:
    df_asset, df_liab = load_data()
    data_loaded = True
except Exception as e:
    st.error(f"数据加载失败: {e}")
    data_loaded = False

if data_loaded:
    # 标题
    st.markdown('<div class="main-header">美联储资产负债表</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="sub-header">数据范围: {df_asset["日期"].min().strftime("%Y-%m-%d")} 至 {df_asset["日期"].max().strftime("%Y-%m-%d")} (周频)</div>', unsafe_allow_html=True)

    # 侧边栏
    with st.sidebar:
        # 侧边栏Logo
        if images_loaded:
            st.markdown(f"""
            <div style="text-align: center; padding: 20px 0; border-bottom: 1px solid rgba(255,255,255,0.1); margin-bottom: 20px;">
                <img src="data:image/png;base64,{logo_base64}" style="width: 120px; filter: brightness(0) invert(1);">
            </div>
            """, unsafe_allow_html=True)

        st.markdown("""
        <div style="color: white; font-size: 15px; font-weight: 600; margin-bottom: 5px;">
            美联储资产负债表
        </div>
        <div style="color: rgba(255,255,255,0.4); font-size: 11px; margin-bottom: 20px; padding-bottom: 15px; border-bottom: 1px solid rgba(255,255,255,0.1);">
            Fed Balance Sheet
        </div>
        """, unsafe_allow_html=True)

        # 固定图表类型为时间序列图
        chart_type = "时间序列图"

        # 使用全部数据
        df_asset_filtered = df_asset.copy()
        df_liab_filtered = df_liab.copy()
        start_date = df_asset['日期'].min()
        end_date = df_asset['日期'].max()

        # 侧边栏底部二维码
        if images_loaded:
            st.markdown("""
            <div style="height: 30px;"></div>
            <div style="border-top: 1px solid rgba(255,255,255,0.1); padding-top: 20px; text-align: center;">
                <div style="color: rgba(255,255,255,0.4); font-size: 10px; margin-bottom: 10px;">关注国信宏观</div>
            </div>
            """, unsafe_allow_html=True)
            st.markdown(f"""
            <div style="text-align: center;">
                <img src="data:image/jpeg;base64,{qrcode_base64}" style="width: 90px; border-radius: 6px;">
            </div>
            """, unsafe_allow_html=True)

    # Tab选择
    tab0, tab1, tab2, tab3, tab4 = st.tabs(["资产负债表", "资产概览", "负债概览", "流动性", "详细数据"])

    # ==================== 资产负债表（交互式） ====================
    with tab0:
        # 日期选择器
        st.markdown('<div class="chart-title">美联储资产负债表</div>', unsafe_allow_html=True)

        # 创建可选日期列表（格式：2026年第20周(2026-05-13)）
        all_dates = df_asset['日期'].sort_values(ascending=False).tolist()
        date_options = []
        for d in all_dates:
            if pd.notna(d):
                week_num = d.isocalendar()[1]
                date_str = d.strftime('%Y-%m-%d')
                date_options.append(f"{d.year}年第{week_num}周({date_str})")

        # 默认选择最新日期
        selected_date_idx = st.selectbox(
            "选择日期查看资产负债表",
            range(len(date_options)),
            format_func=lambda x: date_options[x],
            index=0,
            key='bs_date_selector'
        )

        # 获取选中日期
        selected_date = all_dates[selected_date_idx]

        # 根据选中日期筛选数据
        selected_asset_row = df_asset[df_asset['日期'] == selected_date].iloc[0] if len(df_asset[df_asset['日期'] == selected_date]) > 0 else df_asset.iloc[-1]
        selected_liab_row = df_liab[df_liab['日期'] == selected_date].iloc[0] if len(df_liab[df_liab['日期'] == selected_date]) > 0 else df_liab.iloc[-1]

        latest_asset = selected_asset_row
        latest_liab = selected_liab_row
        total_asset = latest_asset['总资产']
        total_liab = latest_liab['总负债']

        # RGB颜色方案
        colors = ['rgb(4, 68, 119)', 'rgb(153, 204, 255)', 'rgb(216, 12, 24)', 'rgb(94, 94, 94)', 'rgb(148, 148, 149)',
                  'rgb(0, 102, 204)', 'rgb(220, 80, 60)', 'rgb(40, 160, 100)', 'rgb(255, 140, 0)', 'rgb(128, 80, 180)']

        # 初始化默认选中项
        if 'bs_initialized' not in st.session_state:
            st.session_state.bs_initialized = True
            st.session_state['chk_a_持有证券及贷款总额'] = True

        # 定义所有可选项（按顺序）
        all_items = [
            ('asset', '总资产', '总资产'),
            ('asset', '持有证券及贷款总额', '持有证券及贷款总额'),
            ('asset', '持有证券', '持有证券'),
            ('asset', '美国国债', '美国国债'),
            ('asset', '国债-短期债券', '短期债券'),
            ('asset', '国债-中长期名义债券', '中长期名义债券'),
            ('asset', '国债-中长期通胀指数债券', '中长期通胀指数债券'),
            ('asset', '国债-通胀补偿', '通胀补偿'),
            ('asset', '国债-15天以内', '15天以内'),
            ('asset', '国债-16-90天', '16-90天'),
            ('asset', '国债-91天-1年', '91天-1年'),
            ('asset', '国债-1-5年', '1-5年'),
            ('asset', '国债-5-10年', '5-10年'),
            ('asset', '国债-大于10年', '大于10年'),
            ('asset', '联邦机构债券', '联邦机构债券'),
            ('asset', 'MBS', '抵押贷款支持债券(MBS)'),
            ('asset', '未摊销证券溢价', '未摊销证券溢价'),
            ('asset', '未摊销证券折扣', '未摊销证券折扣'),
            ('asset', '正向回购协议', '正向回购协议'),
            ('asset', '其他贷款', '其他贷款'),
            ('asset', '黄金证券账户', '黄金证券账户'),
            ('asset', '特别提款权账户', '特别提款权账户'),
            ('asset', '硬币', '硬币'),
            ('asset', '中央银行流动性互换', '中央银行流动性互换'),
            ('asset', '外币', '外币'),
            ('asset', '其他资产', '其他资产'),
            ('liab', '总负债', '总负债'),
            ('liab', '存款', '存款'),
            ('liab', '存款-定期存款', '存款机构定期存款'),
            ('liab', '存款-其他存款', '存款机构其他存款'),
            ('liab', '财政部一般账户', '美国财政部一般账户'),
            ('liab', '外国官方存款', '外国官方'),
            ('liab', '其他存款', '其他存款'),
            ('liab', '逆向回购协议', '逆向回购协议'),
            ('liab', '联储票据', '联储票据'),
            ('liab', '延迟入账现金', '延迟入账现金项目'),
            ('liab', '其他负债', '其他负债及应计股息'),
            ('liab', '实缴资本', '实缴资本'),
            ('liab', '资本结余', '结余'),
            ('liab', '净资产', '净资产'),
        ]

        # 从session state收集选中项（包含轴信息）
        selected_items = []
        for df_type, col_name, display_name in all_items:
            key = f"chk_a_{col_name}" if df_type == 'asset' else f"chk_l_{col_name}"
            axis_key = f"axis_{col_name}"
            if st.session_state.get(key, False):
                axis = st.session_state.get(axis_key, 'left')
                selected_items.append((df_type, col_name, display_name, axis))

        # ========== 图表控制面板 ==========
        st.markdown("**图表设置**")
        ctrl_col1, ctrl_col2, ctrl_col3, ctrl_col4 = st.columns(4)
        with ctrl_col1:
            transform_type = st.selectbox(
                "数据变换",
                ["raw", "yoy", "mom", "log"],
                format_func=lambda x: {"raw": "原始值", "yoy": "同比变化", "mom": "环比变化", "log": "对数值"}[x],
                key='bs_transform'
            )
        with ctrl_col2:
            resample_freq = st.selectbox(
                "变频",
                ["raw", "M", "Q", "A"],
                format_func=lambda x: {"raw": "原始(周度)", "M": "月度", "Q": "季度", "A": "年度"}[x],
                key='bs_resample'
            )
        with ctrl_col3:
            agg_method = st.selectbox(
                "聚合方式",
                ["mean", "sum", "last"],
                format_func=lambda x: {"mean": "平均值", "sum": "加总值", "last": "期末值"}[x],
                key='bs_agg'
            )
        with ctrl_col4:
            y_scale = st.selectbox(
                "Y轴刻度",
                ["linear", "log"],
                format_func=lambda x: {"linear": "线性", "log": "对数"}[x],
                key='bs_yscale'
            )

        # ========== 时间序列图（上边） ==========
        st.markdown(f"<p style='color: #666; font-size: 12px; margin-bottom: 8px;'>数据日期: {selected_date.strftime('%Y-%m-%d')} | 勾选项目显示时间序列 | 点击「轴」切换左/右轴</p>", unsafe_allow_html=True)

        # 筛选到选中日期为止的数据用于图表
        df_asset_chart = df_asset[df_asset['日期'] <= selected_date]
        df_liab_chart = df_liab[df_liab['日期'] <= selected_date]

        if selected_items:
            fig = go.Figure()
            has_right_axis = any(item[3] == 'right' for item in selected_items)

            for i, (df_type, col_name, display_name, axis) in enumerate(selected_items):
                df = df_asset_chart if df_type == 'asset' else df_liab_chart
                if col_name in df.columns:
                    # 获取原始数据
                    dates = df['日期'].tolist()
                    values = df[col_name].tolist()

                    # 应用变频
                    if resample_freq != 'raw':
                        dates, values = resample_data(dates, values, resample_freq, agg_method)

                    # 应用变换
                    dates, values = apply_transform(dates, values, transform_type)

                    # 确定Y轴
                    yaxis = 'y2' if axis == 'right' else 'y'

                    # 根据变换类型确定数值格式
                    if transform_type == 'yoy':
                        hovertemplate = f'{display_name}: %{{y:.2f}}%<extra></extra>'
                    elif transform_type == 'mom':
                        hovertemplate = f'{display_name}: %{{y:.2f}}%<extra></extra>'
                    elif transform_type == 'log':
                        hovertemplate = f'{display_name}: %{{y:.3f}}<extra></extra>'
                    else:
                        hovertemplate = f'{display_name}: %{{y:,.0f}}<extra></extra>'

                    fig.add_trace(go.Scatter(
                        x=dates,
                        y=values,
                        mode='lines',
                        name=display_name,
                        line=dict(color=colors[i % len(colors)], width=2),
                        hovertemplate=hovertemplate,
                        yaxis=yaxis
                    ))

            # 构建布局
            layout = create_chart_layout("选中项目时间序列")
            layout['yaxis']['type'] = y_scale

            if has_right_axis:
                layout['yaxis2'] = {
                    'title': {'text': '数值 (右轴)', 'font': {'family': 'SimHei, sans-serif', 'size': 12}},
                    'tickfont': {'family': 'SimHei, sans-serif', 'size': 11},
                    'linecolor': '#333',
                    'linewidth': 0.5,
                    'ticks': 'inside',
                    'overlaying': 'y',
                    'side': 'right',
                    'type': y_scale
                }

            fig.update_layout(layout)
            st.plotly_chart(fig, width="stretch")
        else:
            st.info("在下方勾选项目查看时间序列线")

        st.markdown("---")

        # ========== 资产负债表（中间） ==========
        # 辅助函数：渲染行（增加轴选择）
        def render_row(prefix, display_name, value, total, col_name, key_prefix, df_type, gray=False, bold_underline=False, default_checked=False):
            pct = (value / total * 100) if total and total != 0 else 0
            text_color = '#888' if gray else '#333'
            text_style = 'text-decoration: underline; font-weight: 700;' if bold_underline else ''
            key = f"{key_prefix}_{col_name}"
            axis_key = f"axis_{col_name}"

            c1, c2, c3, c4, c5 = st.columns([0.3, 2.5, 1.2, 0.7, 0.5])
            with c1:
                is_selected = st.checkbox("", key=key, label_visibility="collapsed", value=default_checked)
            with c2:
                st.markdown(f"<span style='font-size: 13px; color: {text_color}; {text_style}'>{prefix}{display_name}</span>", unsafe_allow_html=True)
            with c3:
                st.markdown(f"<span style='font-size: 13px; color: {text_color};'>{value:,.0f}</span>", unsafe_allow_html=True)
            with c4:
                st.markdown(f"<span style='font-size: 12px; color: #888;'>{pct:.1f}%</span>", unsafe_allow_html=True)
            with c5:
                # 轴选择按钮（简化样式）
                current_axis = st.session_state.get(axis_key, 'left')
                axis_label = '右' if current_axis == 'right' else '左'
                if st.button(axis_label, key=f"btn_{axis_key}", help="点击切换左/右轴", type='secondary'):
                    st.session_state[axis_key] = 'right' if current_axis == 'left' else 'left'
                    st.rerun()

        # 创建两列布局
        col_left, col_right = st.columns(2)

        # ========== 资产列 ==========
        with col_left:
            st.markdown("""
            <div style='font-weight: 600; font-size: 13px; margin-bottom: 4px; padding-bottom: 4px; border-bottom: 1px solid #ddd;'>
                <span style='display: inline-block; width: 50%;'>资产</span>
                <span style='display: inline-block; width: 22%; text-align: right;'>规模(百万美元)</span>
                <span style='display: inline-block; width: 15%; text-align: right;'>占比(%)</span>
                <span style='display: inline-block; width: 10%; text-align: center;'>轴</span>
            </div>
            """, unsafe_allow_html=True)

            render_row("", "总资产", total_asset, total_asset, '总资产', 'chk_a', 'asset', bold_underline=True)
            render_row("* ", "持有证券及贷款总额", latest_asset['持有证券及贷款总额'], total_asset, '持有证券及贷款总额', 'chk_a', 'asset', default_checked=True)
            render_row("　　** ", "持有证券", latest_asset['持有证券'], total_asset, '持有证券', 'chk_a', 'asset')
            render_row("　　　　*** ", "美国国债", latest_asset['美国国债'], total_asset, '美国国债', 'chk_a', 'asset')

            st.markdown("<div style='color: #888; font-size: 11px; margin-left: 60px;'>按类型：</div>", unsafe_allow_html=True)
            render_row("　　　　　　", "短期债券", latest_asset.get('国债-短期债券', 0) or 0, total_asset, '国债-短期债券', 'chk_a', 'asset', gray=True)
            render_row("　　　　　　", "中长期名义债券", latest_asset.get('国债-中长期名义债券', 0) or 0, total_asset, '国债-中长期名义债券', 'chk_a', 'asset', gray=True)
            render_row("　　　　　　", "中长期通胀指数债券", latest_asset.get('国债-中长期通胀指数债券', 0) or 0, total_asset, '国债-中长期通胀指数债券', 'chk_a', 'asset', gray=True)
            render_row("　　　　　　", "通胀补偿", latest_asset.get('国债-通胀补偿', 0) or 0, total_asset, '国债-通胀补偿', 'chk_a', 'asset', gray=True)

            st.markdown("<div style='color: #888; font-size: 11px; margin-left: 60px;'>按期限：</div>", unsafe_allow_html=True)
            render_row("　　　　　　", "15天以内", latest_asset.get('国债-15天以内', 0) or 0, total_asset, '国债-15天以内', 'chk_a', 'asset', gray=True)
            render_row("　　　　　　", "16-90天", latest_asset.get('国债-16-90天', 0) or 0, total_asset, '国债-16-90天', 'chk_a', 'asset', gray=True)
            render_row("　　　　　　", "91天-1年", latest_asset.get('国债-91天-1年', 0) or 0, total_asset, '国债-91天-1年', 'chk_a', 'asset', gray=True)
            render_row("　　　　　　", "1-5年", latest_asset.get('国债-1-5年', 0) or 0, total_asset, '国债-1-5年', 'chk_a', 'asset', gray=True)
            render_row("　　　　　　", "5-10年", latest_asset.get('国债-5-10年', 0) or 0, total_asset, '国债-5-10年', 'chk_a', 'asset', gray=True)
            render_row("　　　　　　", "大于10年", latest_asset.get('国债-大于10年', 0) or 0, total_asset, '国债-大于10年', 'chk_a', 'asset', gray=True)

            render_row("　　　　*** ", "联邦机构债券", latest_asset.get('联邦机构债券', 0) or 0, total_asset, '联邦机构债券', 'chk_a', 'asset')
            render_row("　　　　*** ", "抵押贷款支持债券(MBS)", latest_asset.get('MBS', 0) or 0, total_asset, 'MBS', 'chk_a', 'asset')
            render_row("　　　　*** ", "未摊销证券溢价", latest_asset.get('未摊销证券溢价', 0) or 0, total_asset, '未摊销证券溢价', 'chk_a', 'asset')
            render_row("　　　　*** ", "未摊销证券折扣", latest_asset.get('未摊销证券折扣', 0) or 0, total_asset, '未摊销证券折扣', 'chk_a', 'asset')
            render_row("　　** ", "正向回购协议", latest_asset.get('正向回购协议', 0) or 0, total_asset, '正向回购协议', 'chk_a', 'asset')
            render_row("　　** ", "其他贷款", latest_asset.get('其他贷款', 0) or 0, total_asset, '其他贷款', 'chk_a', 'asset')
            render_row("* ", "黄金证券账户", latest_asset.get('黄金证券账户', 0) or 0, total_asset, '黄金证券账户', 'chk_a', 'asset')
            render_row("* ", "特别提款权账户", latest_asset.get('特别提款权账户', 0) or 0, total_asset, '特别提款权账户', 'chk_a', 'asset')
            render_row("* ", "硬币", latest_asset.get('硬币', 0) or 0, total_asset, '硬币', 'chk_a', 'asset')
            render_row("* ", "中央银行流动性互换", latest_asset.get('中央银行流动性互换', 0) or 0, total_asset, '中央银行流动性互换', 'chk_a', 'asset')
            render_row("* ", "外币", latest_asset.get('外币', 0) or 0, total_asset, '外币', 'chk_a', 'asset')
            render_row("* ", "其他资产", latest_asset.get('其他资产', 0) or 0, total_asset, '其他资产', 'chk_a', 'asset')

        # ========== 负债列 ==========
        with col_right:
            st.markdown("""
            <div style='font-weight: 600; font-size: 13px; margin-bottom: 4px; padding-bottom: 4px; border-bottom: 1px solid #ddd;'>
                <span style='display: inline-block; width: 50%;'>负债</span>
                <span style='display: inline-block; width: 22%; text-align: right;'>规模(百万美元)</span>
                <span style='display: inline-block; width: 15%; text-align: right;'>占比(%)</span>
                <span style='display: inline-block; width: 10%; text-align: center;'>轴</span>
            </div>
            """, unsafe_allow_html=True)

            render_row("", "总负债", total_liab, total_liab, '总负债', 'chk_l', 'liab', bold_underline=True)
            render_row("* ", "存款", latest_liab['存款'], total_liab, '存款', 'chk_l', 'liab')
            render_row("　　** ", "存款机构定期存款", latest_liab.get('存款-定期存款', 0) or 0, total_liab, '存款-定期存款', 'chk_l', 'liab')
            render_row("　　** ", "存款机构其他存款", latest_liab.get('存款-其他存款', 0) or 0, total_liab, '存款-其他存款', 'chk_l', 'liab')
            render_row("　　　　*** ", "美国财政部一般账户", latest_liab.get('财政部一般账户', 0) or 0, total_liab, '财政部一般账户', 'chk_l', 'liab')
            render_row("　　　　*** ", "外国官方", latest_liab.get('外国官方存款', 0) or 0, total_liab, '外国官方存款', 'chk_l', 'liab')
            render_row("　　　　*** ", "其他存款", latest_liab.get('其他存款', 0) or 0, total_liab, '其他存款', 'chk_l', 'liab')
            render_row("* ", "逆向回购协议", latest_liab.get('逆向回购协议', 0) or 0, total_liab, '逆向回购协议', 'chk_l', 'liab')
            render_row("* ", "联储票据", latest_liab.get('联储票据', 0) or 0, total_liab, '联储票据', 'chk_l', 'liab')
            render_row("* ", "延迟入账现金项目", latest_liab.get('延迟入账现金', 0) or 0, total_liab, '延迟入账现金', 'chk_l', 'liab')
            render_row("* ", "其他负债及应计股息", latest_liab.get('其他负债', 0) or 0, total_liab, '其他负债', 'chk_l', 'liab')

            st.markdown("<div style='color: #333; font-size: 13px; margin: 6px 0 2px 0;'>* 资本账户</div>", unsafe_allow_html=True)
            render_row("　　** ", "实缴资本", latest_liab.get('实缴资本', 0) or 0, total_liab, '实缴资本', 'chk_l', 'liab')
            render_row("　　** ", "结余", latest_liab.get('资本结余', 0) or 0, total_liab, '资本结余', 'chk_l', 'liab')

            st.markdown("<div style='border-top: 1px solid #ddd; margin: 6px 0;'></div>", unsafe_allow_html=True)
            render_row("", "净资产", latest_liab.get('净资产', 0) or 0, total_liab, '净资产', 'chk_l', 'liab', bold_underline=True)

        # ========== Logo和二维码（下边） ==========
        st.markdown("---")
        if images_loaded:
            st.markdown(f"""
            <div style="display: flex; justify-content: center; align-items: center; gap: 30px; padding: 10px;">
                <div style="text-align: center;">
                    <img src="data:image/png;base64,{logo_base64}" style="width: 80px;">
                </div>
                <div style="text-align: center;">
                    <img src="data:image/jpeg;base64,{qrcode_base64}" style="width: 70px; border-radius: 6px;">
                    <p style="color: #888; font-size: 10px; margin: 3px 0 0 0;">关注公众号</p>
                </div>
            </div>
            """, unsafe_allow_html=True)

    # ==================== 资产概览 ====================
    with tab1:
        st.markdown('<div class="chart-title">资产端概览</div>', unsafe_allow_html=True)

        # 日期选择器
        all_dates_asset = df_asset['日期'].sort_values(ascending=False).tolist()
        date_options_asset = []
        for d in all_dates_asset:
            if pd.notna(d):
                week_num = d.isocalendar()[1]
                date_str = d.strftime('%Y-%m-%d')
                date_options_asset.append(f"{d.year}年第{week_num}周({date_str})")

        # 两个日期选择器：当前日期和基期
        col_date1, col_date2 = st.columns(2)
        with col_date1:
            asset_date_idx = st.selectbox(
                "选择日期",
                range(len(date_options_asset)),
                format_func=lambda x: date_options_asset[x],
                index=0,
                key='asset_date_selector'
            )
        with col_date2:
            # 基期选择
            base_period_options = ["上一周", "一个月前", "三个月前", "一年前", "五年前", "十年前", "自定义"]
            base_period_sel = st.selectbox(
                "选择基期",
                base_period_options,
                index=0,
                key='asset_base_period'
            )

        # 如果选择自定义，显示第二个日期选择器
        if base_period_sel == "自定义":
            base_date_idx = st.selectbox(
                "选择基期日期",
                range(len(date_options_asset)),
                format_func=lambda x: date_options_asset[x],
                index=1,  # 默认选择第二个（上一个日期）
                key='asset_base_date_selector'
            )
            base_idx_asset = len(df_asset) - 1 - base_date_idx
        else:
            # 根据选择计算基期索引
            selected_date_asset = all_dates_asset[asset_date_idx]
            selected_idx_asset = df_asset[df_asset['日期'] == selected_date_asset].index[0] if len(df_asset[df_asset['日期'] == selected_date_asset]) > 0 else len(df_asset) - 1

            if base_period_sel == "上一周":
                base_idx_asset = selected_idx_asset - 1 if selected_idx_asset > 0 else selected_idx_asset
            elif base_period_sel == "一个月前":
                # 找约4周前的数据
                base_idx_asset = max(0, selected_idx_asset - 4)
            elif base_period_sel == "三个月前":
                base_idx_asset = max(0, selected_idx_asset - 13)
            elif base_period_sel == "一年前":
                base_idx_asset = max(0, selected_idx_asset - 52)
            elif base_period_sel == "五年前":
                base_idx_asset = max(0, selected_idx_asset - 260)
            elif base_period_sel == "十年前":
                base_idx_asset = max(0, selected_idx_asset - 520)
            else:
                base_idx_asset = max(0, selected_idx_asset - 1)

        selected_date_asset = all_dates_asset[asset_date_idx]
        selected_asset_row = df_asset[df_asset['日期'] == selected_date_asset].iloc[0] if len(df_asset[df_asset['日期'] == selected_date_asset]) > 0 else df_asset.iloc[-1]

        # 获取基期数据
        selected_idx_asset = df_asset[df_asset['日期'] == selected_date_asset].index[0] if len(df_asset[df_asset['日期'] == selected_date_asset]) > 0 else len(df_asset) - 1
        if base_period_sel == "自定义":
            base_idx_asset = len(df_asset) - 1 - base_date_idx
        prev_asset_row = df_asset.iloc[base_idx_asset]

        # 显示对比时间段
        base_date_display = df_asset.iloc[base_idx_asset]['日期'].strftime('%Y-%m-%d')
        st.markdown(f"<div style='color: #666; font-size: 12px; margin-bottom: 10px;'>对比基期: {base_date_display} ({base_period_sel})</div>", unsafe_allow_html=True)

        # 关键指标卡片
        latest_asset = selected_asset_row
        prev_asset = prev_asset_row

        # 第一行：主要资产指标
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            total_asset = latest_asset['总资产']
            change = total_asset - prev_asset['总资产']
            change_pct = change / prev_asset['总资产'] * 100 if prev_asset['总资产'] != 0 else 0
            st.metric("总资产 (百万美元)", f"{total_asset:,.0f}", f"{change:+,.0f} ({change_pct:+.2f}%)")

        with col2:
            securities = latest_asset['持有证券']
            prev_securities = prev_asset.get('持有证券', 0) or 0
            change_sec = securities - prev_securities
            change_sec_pct = change_sec / prev_securities * 100 if prev_securities != 0 else 0
            st.metric("持有证券", f"{securities:,.0f}", f"{change_sec:+,.0f} ({change_sec_pct:+.2f}%)")

        with col3:
            treasuries = latest_asset['美国国债']
            prev_treasuries = prev_asset.get('美国国债', 0) or 0
            change_treas = treasuries - prev_treasuries
            change_treas_pct = change_treas / prev_treasuries * 100 if prev_treasuries != 0 else 0
            st.metric("美国国债", f"{treasuries:,.0f}", f"{change_treas:+,.0f} ({change_treas_pct:+.2f}%)")

        with col4:
            mbs = latest_asset.get('MBS', 0) or 0
            prev_mbs = prev_asset.get('MBS', 0) or 0
            change_mbs = mbs - prev_mbs
            change_mbs_pct = change_mbs / prev_mbs * 100 if prev_mbs != 0 else 0
            st.metric("MBS", f"{mbs:,.0f}", f"{change_mbs:+,.0f} ({change_mbs_pct:+.2f}%)")

        # 第二行：国债按期限
        st.markdown("<div style='color: #666; font-size: 12px; margin-top: 10px;'>美国国债按期限</div>", unsafe_allow_html=True)
        col5, col6, col7, col8 = st.columns(4)
        with col5:
            val_15d = latest_asset.get('国债-15天以内', 0) or 0
            prev_15d = prev_asset.get('国债-15天以内', 0) or 0
            change_15d = val_15d - prev_15d
            pct_15d = change_15d / prev_15d * 100 if prev_15d != 0 else 0
            st.metric("15天以内", f"{val_15d:,.0f}", f"{change_15d:+,.0f} ({pct_15d:+.2f}%)")

        with col6:
            val_90d = latest_asset.get('国债-16-90天', 0) or 0
            prev_90d = prev_asset.get('国债-16-90天', 0) or 0
            change_90d = val_90d - prev_90d
            pct_90d = change_90d / prev_90d * 100 if prev_90d != 0 else 0
            st.metric("16-90天", f"{val_90d:,.0f}", f"{change_90d:+,.0f} ({pct_90d:+.2f}%)")

        with col7:
            val_1y = latest_asset.get('国债-91天-1年', 0) or 0
            prev_1y = prev_asset.get('国债-91天-1年', 0) or 0
            change_1y = val_1y - prev_1y
            pct_1y = change_1y / prev_1y * 100 if prev_1y != 0 else 0
            st.metric("91天-1年", f"{val_1y:,.0f}", f"{change_1y:+,.0f} ({pct_1y:+.2f}%)")

        with col8:
            val_5y = latest_asset.get('国债-1-5年', 0) or 0
            prev_5y = prev_asset.get('国债-1-5年', 0) or 0
            change_5y = val_5y - prev_5y
            pct_5y = change_5y / prev_5y * 100 if prev_5y != 0 else 0
            st.metric("1-5年", f"{val_5y:,.0f}", f"{change_5y:+,.0f} ({pct_5y:+.2f}%)")

        col9, col10, col11, col12 = st.columns(4)
        with col9:
            val_10y = latest_asset.get('国债-5-10年', 0) or 0
            prev_10y = prev_asset.get('国债-5-10年', 0) or 0
            change_10y = val_10y - prev_10y
            pct_10y = change_10y / prev_10y * 100 if prev_10y != 0 else 0
            st.metric("5-10年", f"{val_10y:,.0f}", f"{change_10y:+,.0f} ({pct_10y:+.2f}%)")

        with col10:
            val_g10 = latest_asset.get('国债-大于10年', 0) or 0
            prev_g10 = prev_asset.get('国债-大于10年', 0) or 0
            change_g10 = val_g10 - prev_g10
            pct_g10 = change_g10 / prev_g10 * 100 if prev_g10 != 0 else 0
            st.metric("大于10年", f"{val_g10:,.0f}", f"{change_g10:+,.0f} ({pct_g10:+.2f}%)")

        with col11:
            val_short = latest_asset.get('国债-短期债券', 0) or 0
            prev_short = prev_asset.get('国债-短期债券', 0) or 0
            change_short = val_short - prev_short
            pct_short = change_short / prev_short * 100 if prev_short != 0 else 0
            st.metric("短期债券", f"{val_short:,.0f}", f"{change_short:+,.0f} ({pct_short:+.2f}%)")

        with col12:
            val_nominal = latest_asset.get('国债-中长期名义债券', 0) or 0
            prev_nominal = prev_asset.get('国债-中长期名义债券', 0) or 0
            change_nominal = val_nominal - prev_nominal
            pct_nominal = change_nominal / prev_nominal * 100 if prev_nominal != 0 else 0
            st.metric("中长期名义债券", f"{val_nominal:,.0f}", f"{change_nominal:+,.0f} ({pct_nominal:+.2f}%)")

        st.markdown("---")

        # 国债展示方式选择
        treasury_view = st.radio("美国国债展示方式", ["按类型", "按期限"], horizontal=True, key='treasury_view')

        # 定义饼图颜色
        pie_colors = ['rgb(4, 68, 119)', 'rgb(153, 204, 255)', 'rgb(216, 12, 24)', 'rgb(94, 94, 94)',
                      'rgb(148, 148, 149)', 'rgb(0, 102, 204)', 'rgb(220, 80, 60)', 'rgb(40, 160, 100)']

        # 获取不同时间点的数据
        def get_asset_structure_data(df, idx, view_type):
            """获取资产结构数据"""
            if idx >= len(df):
                idx = len(df) - 1
            row = df.iloc[idx]
            total = row['总资产']

            if view_type == "按类型":
                # 按类型划分国债
                return {
                    '美国国债-短期债券': row.get('国债-短期债券', 0) or 0,
                    '美国国债-中长期名义债券': row.get('国债-中长期名义债券', 0) or 0,
                    '美国国债-通胀指数债券': row.get('国债-中长期通胀指数债券', 0) or 0,
                    '美国国债-通胀补偿': row.get('国债-通胀补偿', 0) or 0,
                    'MBS': row.get('MBS', 0) or 0,
                    '联邦机构债券': row.get('联邦机构债券', 0) or 0,
                    '正向回购协议': row.get('正向回购协议', 0) or 0,
                    '其他资产': (row.get('其他资产', 0) or 0) + (row.get('外币', 0) or 0) +
                              (row.get('黄金证券账户', 0) or 0) + (row.get('特别提款权账户', 0) or 0) +
                              (row.get('硬币', 0) or 0) + (row.get('中央银行流动性互换', 0) or 0) +
                              (row.get('其他贷款', 0) or 0) + (row.get('未摊销证券溢价', 0) or 0) +
                              (row.get('未摊销证券折扣', 0) or 0)
                }
            else:
                # 按期限划分国债
                return {
                    '国债-15天以内': row.get('国债-15天以内', 0) or 0,
                    '国债-16-90天': row.get('国债-16-90天', 0) or 0,
                    '国债-91天-1年': row.get('国债-91天-1年', 0) or 0,
                    '国债-1-5年': row.get('国债-1-5年', 0) or 0,
                    '国债-5-10年': row.get('国债-5-10年', 0) or 0,
                    '国债-大于10年': row.get('国债-大于10年', 0) or 0,
                    'MBS': row.get('MBS', 0) or 0,
                    '联邦机构债券': row.get('联邦机构债券', 0) or 0,
                    '正向回购协议': row.get('正向回购协议', 0) or 0,
                    '其他资产': (row.get('其他资产', 0) or 0) + (row.get('外币', 0) or 0) +
                              (row.get('黄金证券账户', 0) or 0) + (row.get('特别提款权账户', 0) or 0) +
                              (row.get('硬币', 0) or 0) + (row.get('中央银行流动性互换', 0) or 0) +
                              (row.get('其他贷款', 0) or 0) + (row.get('未摊销证券溢价', 0) or 0) +
                              (row.get('未摊销证券折扣', 0) or 0)
                }

        def create_pie_chart(data, title, colors):
            """创建饼图"""
            fig = go.Figure()
            labels = list(data.keys())
            values = [abs(v) for v in data.values()]
            # 过滤掉值为0的项
            non_zero = [(l, v) for l, v in zip(labels, values) if v > 0]
            if non_zero:
                labels, values = zip(*non_zero)
            else:
                labels, values = ['无数据'], [1]

            fig.add_trace(go.Pie(
                labels=labels,
                values=values,
                hole=0.4,
                marker_colors=colors[:len(labels)],
                textinfo='percent',
                textposition='inside',
                hovertemplate='%{label}: %{value:,.0f} (%{percent})<extra></extra>'
            ))
            fig.update_layout(
                title=dict(text=title, font=dict(size=12)),
                showlegend=True,
                legend=dict(orientation='h', yanchor='bottom', y=-0.2, xanchor='center', x=0.5, font=dict(size=7)),
                margin=dict(t=30, b=40, l=10, r=10),
                height=280
            )
            return fig

        # 计算6个时间点
        total_len = len(df_asset_filtered)
        idx_current = total_len - 1
        idx_1w = max(0, total_len - 2)  # 1周前
        idx_1m = max(0, total_len - 5)  # 1个月前（约4-5周）
        idx_3m = max(0, total_len - 13)  # 3个月前（约13周）
        idx_1y = max(0, total_len - 52)  # 1年前（约52周）
        idx_5y = max(0, total_len - 260)  # 5年前（约260周）

        dates = df_asset_filtered['日期'].tolist()
        date_current = dates[idx_current].strftime('%Y-%m-%d')
        date_1w = dates[idx_1w].strftime('%Y-%m-%d')
        date_1m = dates[idx_1m].strftime('%Y-%m-%d')
        date_3m = dates[idx_3m].strftime('%Y-%m-%d')
        date_1y = dates[idx_1y].strftime('%Y-%m-%d')
        date_5y = dates[idx_5y].strftime('%Y-%m-%d')

        # 3x2 环状图
        st.markdown("#### 资产结构对比")
        col_a, col_b = st.columns(2)
        col_c, col_d = st.columns(2)
        col_e, col_f = st.columns(2)

        with col_a:
            data_current = get_asset_structure_data(df_asset_filtered, idx_current, treasury_view)
            st.plotly_chart(create_pie_chart(data_current, f"当前 ({date_current})", pie_colors), width="stretch")

        with col_b:
            data_1w = get_asset_structure_data(df_asset_filtered, idx_1w, treasury_view)
            st.plotly_chart(create_pie_chart(data_1w, f"1周前 ({date_1w})", pie_colors), width="stretch")

        with col_c:
            data_1m = get_asset_structure_data(df_asset_filtered, idx_1m, treasury_view)
            st.plotly_chart(create_pie_chart(data_1m, f"1个月前 ({date_1m})", pie_colors), width="stretch")

        with col_d:
            data_3m = get_asset_structure_data(df_asset_filtered, idx_3m, treasury_view)
            st.plotly_chart(create_pie_chart(data_3m, f"3个月前 ({date_3m})", pie_colors), width="stretch")

        with col_e:
            data_1y = get_asset_structure_data(df_asset_filtered, idx_1y, treasury_view)
            st.plotly_chart(create_pie_chart(data_1y, f"1年前 ({date_1y})", pie_colors), width="stretch")

        with col_f:
            data_5y = get_asset_structure_data(df_asset_filtered, idx_5y, treasury_view)
            st.plotly_chart(create_pie_chart(data_5y, f"5年前 ({date_5y})", pie_colors), width="stretch")

        st.markdown("---")

        # 边际变化表格
        st.markdown("#### 资产边际变化")

        # 定义要展示的资产项（带层级）
        # (列名, 显示名, 层级) 层级: 0=总, 1=一级, 2=二级, 3=三级, 4=四级
        asset_items = [
            # 总资产
            ('总资产', '总资产', 0),

            # * 持有证券及贷款总额
            ('持有证券及贷款总额', '* 持有证券及贷款总额', 1),

            # ** 持有证券
            ('持有证券', '　　** 持有证券', 2),

            # *** 美国国债
            ('美国国债', '　　　　*** 美国国债', 3),

            # 按类型 - 四级
            ('国债-短期债券', '　　　　　　短期债券', 4),
            ('国债-中长期名义债券', '　　　　　　中长期名义债券', 4),
            ('国债-中长期通胀指数债券', '　　　　　　中长期通胀指数债券', 4),
            ('国债-通胀补偿', '　　　　　　通胀补偿', 4),

            # 按期限 - 四级
            ('国债-15天以内', '　　　　　　15天以内', 4),
            ('国债-16-90天', '　　　　　　16-90天', 4),
            ('国债-91天-1年', '　　　　　　91天-1年', 4),
            ('国债-1-5年', '　　　　　　1-5年', 4),
            ('国债-5-10年', '　　　　　　5-10年', 4),
            ('国债-大于10年', '　　　　　　大于10年', 4),

            # *** 联邦机构债券
            ('联邦机构债券', '　　　　*** 联邦机构债券', 3),

            # *** MBS
            ('MBS', '　　　　*** 抵押贷款支持债券(MBS)', 3),

            # *** 未摊销证券溢价/折扣
            ('未摊销证券溢价', '　　　　*** 未摊销证券溢价', 3),
            ('未摊销证券折扣', '　　　　*** 未摊销证券折扣', 3),

            # ** 正向回购协议
            ('正向回购协议', '　　** 正向回购协议', 2),

            # ** 其他贷款
            ('其他贷款', '　　** 其他贷款', 2),

            # * 其他一级项目
            ('黄金证券账户', '* 黄金证券账户', 1),
            ('特别提款权账户', '* 特别提款权账户', 1),
            ('硬币', '* 硬币', 1),
            ('中央银行流动性互换', '* 中央银行流动性互换', 1),
            ('外币', '* 外币', 1),

            # 其他资产
            ('托收中项目', '* 托收中项目', 1),
            ('银行不动产', '* 银行不动产', 1),
            ('MSF投资组合', '* MSF投资组合', 1),
            ('市政流动性基金', '* 市政流动性基金', 1),
            ('TALF II投资组合', '* TALF II投资组合', 1),

            ('其他资产', '* 其他资产', 1),
        ]

        # 定义时间点（周为单位）
        time_points = [
            ('当周', 0),
            ('1周前', 1),
            ('1个月前', 4),
            ('3个月前', 13),
            ('1年前', 52),
            ('5年前', 260),
            ('10年前', 520),
        ]

        # 获取各时间点的数据
        period_data = []
        period_labels = []
        for label, weeks_ago in time_points:
            idx = total_len - 1 - weeks_ago
            if idx >= 0:
                period_data.append(df_asset_filtered.iloc[idx])
                period_labels.append(f"{label} ({dates[idx].strftime('%Y-%m-%d')})")
            else:
                period_data.append(None)
                period_labels.append(f"{label} (无数据)")

        # 构建表格数据
        table_data = []
        for item in asset_items:
            col_name, display_name, level = item
            row = {'资产项目': display_name}
            for i, (period_df, period_label) in enumerate(zip(period_data, period_labels)):
                if period_df is not None:
                    val = period_df.get(col_name, 0) or 0
                    total = period_df['总资产']
                    pct = (val / total * 100) if total != 0 else 0
                    row[period_label] = f"{val:,.0f} ({pct:.1f}%)"
                else:
                    row[period_label] = "-"
            table_data.append(row)

        # 显示表格
        if table_data:
            df_table = pd.DataFrame(table_data)
            st.dataframe(df_table, use_container_width=True, hide_index=True)

        # ========== 时间趋势柱状图 ==========
        st.markdown("#### ①资产绝对相对规模")

        # 资产选择下拉框
        asset_select_options = [(col, display) for col, display, level in asset_items if col in df_asset_filtered.columns]
        trend_selected_idx = st.selectbox(
            "选择资产项目",
            range(len(asset_select_options)),
            format_func=lambda x: asset_select_options[x][1],
            key='trend_asset_select'
        )
        trend_col = asset_select_options[trend_selected_idx][0]
        trend_display = asset_select_options[trend_selected_idx][1]

        # 获取全口径周度数据
        full_dates = df_asset_filtered['日期'].tolist()
        full_values = (df_asset_filtered[trend_col].fillna(0)).tolist()

        # 计算关键值
        last_val = full_values[-1]
        last_date = full_dates[-1]
        max_val = max(full_values)
        max_idx = full_values.index(max_val)
        max_date = full_dates[max_idx]
        min_val = min(full_values)
        min_idx = full_values.index(min_val)
        min_date = full_dates[min_idx]

        # 两列布局：左侧柱状图，右侧面积图
        col_bar, col_area = st.columns(2)

        with col_bar:
            # 柱状图
            fig_bar = go.Figure()
            fig_bar.add_trace(go.Bar(
                x=full_dates,
                y=full_values,
                marker_color='rgba(4, 68, 119, 0.7)',
                hovertemplate='%{x|%Y-%m-%d}: %{y:,.0f}<extra></extra>'
            ))
            # 标注最后一个点
            fig_bar.add_trace(go.Scatter(
                x=[last_date],
                y=[last_val],
                mode='markers+text',
                marker=dict(size=10, color='rgb(216, 12, 24)'),
                text=[f'{last_val:,.0f}'],
                textposition='top center',
                textfont=dict(size=9, color='rgb(216, 12, 24)'),
                showlegend=False
            ))
            # 标注最高点
            fig_bar.add_trace(go.Scatter(
                x=[max_date],
                y=[max_val],
                mode='markers+text',
                marker=dict(size=8, color='rgb(40, 160, 100)'),
                text=[f'最高: {max_val:,.0f}'],
                textposition='top center',
                textfont=dict(size=8, color='rgb(40, 160, 100)'),
                showlegend=False
            ))
            # 标注最低点
            fig_bar.add_trace(go.Scatter(
                x=[min_date],
                y=[min_val],
                mode='markers+text',
                marker=dict(size=8, color='rgb(94, 94, 94)'),
                text=[f'最低: {min_val:,.0f}'],
                textposition='bottom center',
                textfont=dict(size=8, color='rgb(94, 94, 94)'),
                showlegend=False
            ))
            fig_bar.update_layout(
                title=dict(text=f"{trend_display.strip()} - 柱状图（绝对规模）", font=dict(size=12)),
                xaxis=dict(title='日期'),
                yaxis=dict(title='规模 (百万美元)'),
                showlegend=False,
                margin=dict(t=40, b=30, l=50, r=20),
                height=300
            )
            st.plotly_chart(fig_bar, width="stretch")

        with col_area:
            # 计算占总资产的百分比
            total_asset_values = df_asset_filtered['总资产'].fillna(0).tolist()
            percent_values = [(v / t * 100) if t != 0 else 0 for v, t in zip(full_values, total_asset_values)]
            last_percent = percent_values[-1]

            # 面积图（相对规模）
            fig_area = go.Figure()
            fig_area.add_trace(go.Scatter(
                x=full_dates,
                y=percent_values,
                fill='tozeroy',
                mode='lines',
                line=dict(color='rgb(4, 68, 119)', width=1.5),
                fillcolor='rgba(4, 68, 119, 0.2)',
                hovertemplate='%{x|%Y-%m-%d}: %{y:.2f}%<extra></extra>'
            ))
            # 标注最后一个点
            fig_area.add_trace(go.Scatter(
                x=[last_date],
                y=[last_percent],
                mode='markers+text',
                marker=dict(size=10, color='rgb(216, 12, 24)'),
                text=[f'{last_percent:.2f}%'],
                textposition='top center',
                textfont=dict(size=9, color='rgb(216, 12, 24)'),
                showlegend=False
            ))
            fig_area.update_layout(
                title=dict(text=f"{trend_display.strip()} - 面积图（占总资产%）", font=dict(size=12)),
                xaxis=dict(title='日期'),
                yaxis=dict(title='占比 (%)'),
                showlegend=False,
                margin=dict(t=40, b=30, l=50, r=20),
                height=300
            )
            st.plotly_chart(fig_area, width="stretch")

        st.markdown("---")

        # ========== ②周度变动分析 ==========
        st.markdown("#### ②周度变动分析")

        # 资产选择下拉框
        asset_options = [(col, display) for col, display, level in asset_items if col in df_asset_filtered.columns]
        selected_idx = st.selectbox(
            "选择资产项目",
            range(len(asset_options)),
            format_func=lambda x: asset_options[x][1],
            key='weekly_change_asset'
        )
        selected_asset_col = asset_options[selected_idx][0]
        selected_display = asset_options[selected_idx][1]

        # 计算全口径周度变动
        weekly_changes = []
        weekly_dates = []
        values = df_asset_filtered[selected_asset_col].fillna(0).tolist()
        all_dates = df_asset_filtered['日期'].tolist()

        for i in range(1, len(values)):
            change = values[i] - values[i-1]
            weekly_changes.append(change)
            weekly_dates.append(all_dates[i])

        # 创建柱状图
        fig_bar = go.Figure()
        colors_bar = ['rgb(216, 12, 24)' if v < 0 else 'rgb(4, 68, 119)' for v in weekly_changes]
        fig_bar.add_trace(go.Bar(
            x=weekly_dates,
            y=weekly_changes,
            marker_color=colors_bar,
            hovertemplate='%{x|%Y-%m-%d}: %{y:,.0f}<extra></extra>'
        ))
        # 标注最后一期
        last_change = weekly_changes[-1]
        last_date = weekly_dates[-1]
        fig_bar.add_trace(go.Scatter(
            x=[last_date],
            y=[last_change],
            mode='markers+text',
            marker=dict(size=10, color='rgb(216, 12, 24)' if last_change < 0 else 'rgb(4, 68, 119)'),
            text=[f'{last_change:+,.0f}'],
            textposition='top center' if last_change >= 0 else 'bottom center',
            textfont=dict(size=9, color='rgb(216, 12, 24)' if last_change < 0 else 'rgb(4, 68, 119)'),
            showlegend=False
        ))
        # 标注最大变动
        max_change = max(weekly_changes)
        max_idx = weekly_changes.index(max_change)
        max_date = weekly_dates[max_idx]
        fig_bar.add_trace(go.Scatter(
            x=[max_date],
            y=[max_change],
            mode='markers+text',
            marker=dict(size=8, color='rgb(40, 160, 100)'),
            text=[f'最大增: {max_change:,.0f}'],
            textposition='top center',
            textfont=dict(size=8, color='rgb(40, 160, 100)'),
            showlegend=False
        ))
        # 标注最小变动
        min_change = min(weekly_changes)
        min_idx = weekly_changes.index(min_change)
        min_date = weekly_dates[min_idx]
        fig_bar.add_trace(go.Scatter(
            x=[min_date],
            y=[min_change],
            mode='markers+text',
            marker=dict(size=8, color='rgb(94, 94, 94)'),
            text=[f'最大减: {min_change:,.0f}'],
            textposition='bottom center',
            textfont=dict(size=8, color='rgb(94, 94, 94)'),
            showlegend=False
        ))
        fig_bar.update_layout(
            title=dict(text=f"{selected_display.strip()} 周度变动", font=dict(size=13)),
            xaxis=dict(title='日期'),
            yaxis=dict(title='变动量 (百万美元)'),
            showlegend=False,
            margin=dict(t=40, b=30, l=50, r=20),
            height=350
        )
        st.plotly_chart(fig_bar, width="stretch")

        st.markdown("---")

        # ========== ③环比增速 ==========
        st.markdown("#### ③环比增速")

        # 资产选择下拉框
        growth_select_options = [(col, display) for col, display, level in asset_items if col in df_asset_filtered.columns]
        growth_selected_idx = st.selectbox(
            "选择资产项目",
            range(len(growth_select_options)),
            format_func=lambda x: growth_select_options[x][1],
            key='growth_asset_select'
        )
        growth_col = growth_select_options[growth_selected_idx][0]
        growth_display = growth_select_options[growth_selected_idx][1]

        # 计算环比增速
        values = df_asset_filtered[growth_col].fillna(0).tolist()
        all_dates = df_asset_filtered['日期'].tolist()

        growth_rates = []
        growth_dates = []
        for i in range(1, len(values)):
            if values[i-1] != 0:
                rate = (values[i] - values[i-1]) / abs(values[i-1]) * 100
            else:
                rate = 0 if values[i] == 0 else 100  # 从0增长视为100%
            growth_rates.append(rate)
            growth_dates.append(all_dates[i])

        # 创建环比增速柱状图
        fig_growth = go.Figure()
        colors_growth = ['rgb(216, 12, 24)' if v < 0 else 'rgb(4, 68, 119)' for v in growth_rates]
        fig_growth.add_trace(go.Bar(
            x=growth_dates,
            y=growth_rates,
            marker_color=colors_growth,
            hovertemplate='%{x|%Y-%m-%d}: %{y:.2f}%<extra></extra>'
        ))
        # 标注最后一期
        last_rate = growth_rates[-1]
        last_date = growth_dates[-1]
        fig_growth.add_trace(go.Scatter(
            x=[last_date],
            y=[last_rate],
            mode='markers+text',
            marker=dict(size=10, color='rgb(216, 12, 24)' if last_rate < 0 else 'rgb(4, 68, 119)'),
            text=[f'{last_rate:+.2f}%'],
            textposition='top center' if last_rate >= 0 else 'bottom center',
            textfont=dict(size=9, color='rgb(216, 12, 24)' if last_rate < 0 else 'rgb(4, 68, 119)'),
            showlegend=False
        ))
        # 添加零线
        fig_growth.add_hline(y=0, line_dash='dash', line_color='gray', opacity=0.5)
        fig_growth.update_layout(
            title=dict(text=f"{growth_display.strip()} 环比增速", font=dict(size=13)),
            xaxis=dict(title='日期'),
            yaxis=dict(title='环比增速 (%)'),
            showlegend=False,
            margin=dict(t=40, b=30, l=50, r=20),
            height=300
        )
        st.plotly_chart(fig_growth, width="stretch")

    # ==================== 负债概览 ====================
    with tab2:
        st.markdown('<div class="chart-title">负债端概览</div>', unsafe_allow_html=True)

        # 日期选择器
        all_dates_liab = df_liab['日期'].sort_values(ascending=False).tolist()
        date_options_liab = []
        for d in all_dates_liab:
            if pd.notna(d):
                week_num = d.isocalendar()[1]
                date_str = d.strftime('%Y-%m-%d')
                date_options_liab.append(f"{d.year}年第{week_num}周({date_str})")

        # 两个日期选择器：当前日期和基期
        col_date1, col_date2 = st.columns(2)
        with col_date1:
            liab_date_idx = st.selectbox(
                "选择日期",
                range(len(date_options_liab)),
                format_func=lambda x: date_options_liab[x],
                index=0,
                key='liab_date_selector'
            )
        with col_date2:
            # 基期选择
            base_period_options_liab = ["上一周", "一个月前", "三个月前", "一年前", "五年前", "十年前", "自定义"]
            base_period_sel_liab = st.selectbox(
                "选择基期",
                base_period_options_liab,
                index=0,
                key='liab_base_period'
            )

        # 如果选择自定义，显示第二个日期选择器
        if base_period_sel_liab == "自定义":
            base_date_idx_liab = st.selectbox(
                "选择基期日期",
                range(len(date_options_liab)),
                format_func=lambda x: date_options_liab[x],
                index=1,  # 默认选择第二个（上一个日期）
                key='liab_base_date_selector'
            )
            base_idx_liab = len(df_liab) - 1 - base_date_idx_liab
        else:
            # 根据选择计算基期索引
            selected_date_liab = all_dates_liab[liab_date_idx]
            selected_idx_liab = df_liab[df_liab['日期'] == selected_date_liab].index[0] if len(df_liab[df_liab['日期'] == selected_date_liab]) > 0 else len(df_liab) - 1

            if base_period_sel_liab == "上一周":
                base_idx_liab = selected_idx_liab - 1 if selected_idx_liab > 0 else selected_idx_liab
            elif base_period_sel_liab == "一个月前":
                base_idx_liab = max(0, selected_idx_liab - 4)
            elif base_period_sel_liab == "三个月前":
                base_idx_liab = max(0, selected_idx_liab - 13)
            elif base_period_sel_liab == "一年前":
                base_idx_liab = max(0, selected_idx_liab - 52)
            elif base_period_sel_liab == "五年前":
                base_idx_liab = max(0, selected_idx_liab - 260)
            elif base_period_sel_liab == "十年前":
                base_idx_liab = max(0, selected_idx_liab - 520)
            else:
                base_idx_liab = max(0, selected_idx_liab - 1)

        selected_date_liab = all_dates_liab[liab_date_idx]
        selected_liab_row = df_liab[df_liab['日期'] == selected_date_liab].iloc[0] if len(df_liab[df_liab['日期'] == selected_date_liab]) > 0 else df_liab.iloc[-1]

        # 获取基期数据
        selected_idx_liab = df_liab[df_liab['日期'] == selected_date_liab].index[0] if len(df_liab[df_liab['日期'] == selected_date_liab]) > 0 else len(df_liab) - 1
        if base_period_sel_liab == "自定义":
            base_idx_liab = len(df_liab) - 1 - base_date_idx_liab
        prev_liab_row = df_liab.iloc[base_idx_liab]

        # 显示对比时间段
        base_date_display_liab = df_liab.iloc[base_idx_liab]['日期'].strftime('%Y-%m-%d')
        st.markdown(f"<div style='color: #666; font-size: 12px; margin-bottom: 10px;'>对比基期: {base_date_display_liab} ({base_period_sel_liab})</div>", unsafe_allow_html=True)

        # 关键指标卡片
        latest_liab = selected_liab_row
        prev_liab = prev_liab_row

        # 第一行：主要负债指标
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            total_liab = latest_liab['总负债']
            change = total_liab - prev_liab['总负债']
            change_pct = change / prev_liab['总负债'] * 100 if prev_liab['总负债'] != 0 else 0
            st.metric("总负债 (百万美元)", f"{total_liab:,.0f}", f"{change:+,.0f} ({change_pct:+.2f}%)")

        with col2:
            deposits = latest_liab['存款']
            prev_deposits = prev_liab.get('存款', 0) or 0
            change_dep = deposits - prev_deposits
            change_dep_pct = change_dep / prev_deposits * 100 if prev_deposits != 0 else 0
            st.metric("存款", f"{deposits:,.0f}", f"{change_dep:+,.0f} ({change_dep_pct:+.2f}%)")

        with col3:
            reverse_repo = latest_liab.get('逆向回购协议', 0) or 0
            prev_reverse = prev_liab.get('逆向回购协议', 0) or 0
            change_rrp = reverse_repo - prev_reverse
            change_rrp_pct = change_rrp / prev_reverse * 100 if prev_reverse != 0 else 0
            st.metric("逆向回购协议", f"{reverse_repo:,.0f}", f"{change_rrp:+,.0f} ({change_rrp_pct:+.2f}%)")

        with col4:
            fed_notes = latest_liab.get('联储票据', 0) or 0
            prev_notes = prev_liab.get('联储票据', 0) or 0
            change_notes = fed_notes - prev_notes
            change_notes_pct = change_notes / prev_notes * 100 if prev_notes != 0 else 0
            st.metric("联储票据", f"{fed_notes:,.0f}", f"{change_notes:+,.0f} ({change_notes_pct:+.2f}%)")

        # 第二行：存款明细
        st.markdown("<div style='color: #666; font-size: 12px; margin-top: 10px;'>存款明细</div>", unsafe_allow_html=True)
        col5, col6, col7, col8 = st.columns(4)
        with col5:
            val_tga = latest_liab.get('财政部一般账户', 0) or 0
            prev_tga = prev_liab.get('财政部一般账户', 0) or 0
            change_tga = val_tga - prev_tga
            pct_tga = change_tga / prev_tga * 100 if prev_tga != 0 else 0
            st.metric("财政部一般账户", f"{val_tga:,.0f}", f"{change_tga:+,.0f} ({pct_tga:+.2f}%)")

        with col6:
            val_foreign = latest_liab.get('外国官方存款', 0) or 0
            prev_foreign = prev_liab.get('外国官方存款', 0) or 0
            change_foreign = val_foreign - prev_foreign
            pct_foreign = change_foreign / prev_foreign * 100 if prev_foreign != 0 else 0
            st.metric("外国官方存款", f"{val_foreign:,.0f}", f"{change_foreign:+,.0f} ({pct_foreign:+.2f}%)")

        with col7:
            val_other_dep = latest_liab.get('其他存款', 0) or 0
            prev_other_dep = prev_liab.get('其他存款', 0) or 0
            change_other_dep = val_other_dep - prev_other_dep
            pct_other_dep = change_other_dep / prev_other_dep * 100 if prev_other_dep != 0 else 0
            st.metric("其他存款", f"{val_other_dep:,.0f}", f"{change_other_dep:+,.0f} ({pct_other_dep:+.2f}%)")

        with col8:
            val_time_dep = latest_liab.get('存款-定期存款', 0) or 0
            prev_time_dep = prev_liab.get('存款-定期存款', 0) or 0
            change_time_dep = val_time_dep - prev_time_dep
            pct_time_dep = change_time_dep / prev_time_dep * 100 if prev_time_dep != 0 else 0
            st.metric("定期存款", f"{val_time_dep:,.0f}", f"{change_time_dep:+,.0f} ({pct_time_dep:+.2f}%)")

        # 第三行：资本账户
        st.markdown("<div style='color: #666; font-size: 12px; margin-top: 10px;'>资本账户</div>", unsafe_allow_html=True)
        col9, col10, col11, col12 = st.columns(4)
        with col9:
            val_capital = latest_liab.get('实缴资本', 0) or 0
            prev_capital = prev_liab.get('实缴资本', 0) or 0
            change_capital = val_capital - prev_capital
            st.metric("实缴资本", f"{val_capital:,.0f}", f"{change_capital:+,.0f}")

        with col10:
            val_surplus = latest_liab.get('资本结余', 0) or 0
            prev_surplus = prev_liab.get('资本结余', 0) or 0
            change_surplus = val_surplus - prev_surplus
            st.metric("资本结余", f"{val_surplus:,.0f}", f"{change_surplus:+,.0f}")

        with col11:
            val_net = latest_liab.get('净资产', 0) or 0
            prev_net = prev_liab.get('净资产', 0) or 0
            change_net = val_net - prev_net
            pct_net = change_net / prev_net * 100 if prev_net != 0 else 0
            st.metric("净资产", f"{val_net:,.0f}", f"{change_net:+,.0f} ({pct_net:+.2f}%)")

        with col12:
            val_other_liab = latest_liab.get('其他负债', 0) or 0
            prev_other_liab = prev_liab.get('其他负债', 0) or 0
            change_other_liab = val_other_liab - prev_other_liab
            pct_other_liab = change_other_liab / prev_other_liab * 100 if prev_other_liab != 0 else 0
            st.metric("其他负债", f"{val_other_liab:,.0f}", f"{change_other_liab:+,.0f} ({pct_other_liab:+.2f}%)")

        st.markdown("---")

        # ========== 负债结构对比（环状图） ==========
        pie_colors_liab = ['rgb(4, 68, 119)', 'rgb(153, 204, 255)', 'rgb(216, 12, 24)', 'rgb(94, 94, 94)',
                          'rgb(148, 148, 149)', 'rgb(0, 102, 204)', 'rgb(220, 80, 60)', 'rgb(40, 160, 100)']

        def get_liab_structure_data(df, idx):
            """获取负债结构数据"""
            if idx >= len(df):
                idx = len(df) - 1
            row = df.iloc[idx]
            return {
                '存款': row.get('存款', 0) or 0,
                '逆向回购协议': row.get('逆向回购协议', 0) or 0,
                '联储票据': row.get('联储票据', 0) or 0,
                '其他负债': row.get('其他负债', 0) or 0,
                '实缴资本': row.get('实缴资本', 0) or 0,
                '资本结余': row.get('资本结余', 0) or 0,
            }

        def create_pie_chart_liab(data, title, colors):
            """创建负债饼图"""
            fig = go.Figure()
            labels = list(data.keys())
            values = [abs(v) for v in data.values()]
            non_zero = [(l, v) for l, v in zip(labels, values) if v > 0]
            if non_zero:
                labels, values = zip(*non_zero)
            else:
                labels, values = ['无数据'], [1]
            fig.add_trace(go.Pie(
                labels=labels,
                values=values,
                hole=0.4,
                marker_colors=colors[:len(labels)],
                textinfo='percent',
                textposition='inside',
                hovertemplate='%{label}: %{value:,.0f} (%{percent})<extra></extra>'
            ))
            fig.update_layout(
                title=dict(text=title, font=dict(size=12)),
                showlegend=True,
                legend=dict(orientation='h', yanchor='bottom', y=-0.2, xanchor='center', x=0.5, font=dict(size=7)),
                margin=dict(t=30, b=40, l=10, r=10),
                height=280
            )
            return fig

        # 计算6个时间点
        total_len_liab = len(df_liab_filtered)
        dates_liab = df_liab_filtered['日期'].tolist()
        idx_current_liab = total_len_liab - 1
        idx_1w_liab = max(0, total_len_liab - 2)
        idx_1m_liab = max(0, total_len_liab - 5)
        idx_3m_liab = max(0, total_len_liab - 13)
        idx_1y_liab = max(0, total_len_liab - 52)
        idx_5y_liab = max(0, total_len_liab - 260)

        date_current_liab = dates_liab[idx_current_liab].strftime('%Y-%m-%d')
        date_1w_liab = dates_liab[idx_1w_liab].strftime('%Y-%m-%d')
        date_1m_liab = dates_liab[idx_1m_liab].strftime('%Y-%m-%d')
        date_3m_liab = dates_liab[idx_3m_liab].strftime('%Y-%m-%d')
        date_1y_liab = dates_liab[idx_1y_liab].strftime('%Y-%m-%d')
        date_5y_liab = dates_liab[idx_5y_liab].strftime('%Y-%m-%d')

        st.markdown("#### 负债结构对比")
        col_a, col_b = st.columns(2)
        col_c, col_d = st.columns(2)
        col_e, col_f = st.columns(2)

        with col_a:
            data_current_liab = get_liab_structure_data(df_liab_filtered, idx_current_liab)
            st.plotly_chart(create_pie_chart_liab(data_current_liab, f"当前 ({date_current_liab})", pie_colors_liab), width="stretch")

        with col_b:
            data_1w_liab = get_liab_structure_data(df_liab_filtered, idx_1w_liab)
            st.plotly_chart(create_pie_chart_liab(data_1w_liab, f"1周前 ({date_1w_liab})", pie_colors_liab), width="stretch")

        with col_c:
            data_1m_liab = get_liab_structure_data(df_liab_filtered, idx_1m_liab)
            st.plotly_chart(create_pie_chart_liab(data_1m_liab, f"1个月前 ({date_1m_liab})", pie_colors_liab), width="stretch")

        with col_d:
            data_3m_liab = get_liab_structure_data(df_liab_filtered, idx_3m_liab)
            st.plotly_chart(create_pie_chart_liab(data_3m_liab, f"3个月前 ({date_3m_liab})", pie_colors_liab), width="stretch")

        with col_e:
            data_1y_liab = get_liab_structure_data(df_liab_filtered, idx_1y_liab)
            st.plotly_chart(create_pie_chart_liab(data_1y_liab, f"1年前 ({date_1y_liab})", pie_colors_liab), width="stretch")

        with col_f:
            data_5y_liab = get_liab_structure_data(df_liab_filtered, idx_5y_liab)
            st.plotly_chart(create_pie_chart_liab(data_5y_liab, f"5年前 ({date_5y_liab})", pie_colors_liab), width="stretch")

        st.markdown("---")

        # ========== 负债边际变化表格 ==========
        st.markdown("#### 负债边际变化")

        # 定义要展示的负债项
        liab_items = [
            ('总负债', '总负债', 0),
            ('存款', '* 存款', 1),
            ('存款-定期存款', '　　** 定期存款', 2),
            ('存款-其他存款', '　　** 其他存款', 2),
            ('财政部一般账户', '　　　　*** 财政部一般账户', 3),
            ('外国官方存款', '　　　　*** 外国官方存款', 3),
            ('其他存款', '　　　　*** 其他存款', 3),
            ('逆向回购协议', '* 逆向回购协议', 1),
            ('联储票据', '* 联储票据', 1),
            ('延迟入账现金', '* 延迟入账现金', 1),
            ('其他负债', '* 其他负债', 1),
            ('实缴资本', '* 实缴资本', 1),
            ('资本结余', '* 资本结余', 1),
            ('净资产', '净资产', 0),
        ]

        time_points_liab = [
            ('当周', 0),
            ('1周前', 1),
            ('1个月前', 4),
            ('3个月前', 13),
            ('1年前', 52),
            ('5年前', 260),
            ('10年前', 520),
        ]

        period_data_liab = []
        period_labels_liab = []
        for label, weeks_ago in time_points_liab:
            idx = total_len_liab - 1 - weeks_ago
            if idx >= 0:
                period_data_liab.append(df_liab_filtered.iloc[idx])
                period_labels_liab.append(f"{label} ({dates_liab[idx].strftime('%Y-%m-%d')})")
            else:
                period_data_liab.append(None)
                period_labels_liab.append(f"{label} (无数据)")

        # 构建表格数据
        table_data_liab = []
        for item in liab_items:
            col_name, display_name, level = item
            row = {'负债项目': display_name}
            for i, (period_df, period_label) in enumerate(zip(period_data_liab, period_labels_liab)):
                if period_df is not None:
                    val = period_df.get(col_name, 0) or 0
                    total = period_df['总负债']
                    pct = (val / total * 100) if total != 0 else 0
                    row[period_label] = f"{val:,.0f} ({pct:.1f}%)"
                else:
                    row[period_label] = "-"
            table_data_liab.append(row)

        if table_data_liab:
            df_table_liab = pd.DataFrame(table_data_liab)
            st.dataframe(df_table_liab, use_container_width=True, hide_index=True)

        st.markdown("---")

        # ========== ①负债绝对相对规模 ==========
        st.markdown("#### ①负债绝对相对规模")

        liab_select_options = [(col, display) for col, display, level in liab_items if col in df_liab_filtered.columns]
        trend_liab_idx = st.selectbox(
            "选择负债项目",
            range(len(liab_select_options)),
            format_func=lambda x: liab_select_options[x][1],
            key='trend_liab_select'
        )
        trend_liab_col = liab_select_options[trend_liab_idx][0]
        trend_liab_display = liab_select_options[trend_liab_idx][1]

        full_dates_liab = df_liab_filtered['日期'].tolist()
        full_values_liab = (df_liab_filtered[trend_liab_col].fillna(0)).tolist()

        last_val_liab = full_values_liab[-1]
        last_date_liab = full_dates_liab[-1]
        max_val_liab = max(full_values_liab)
        max_idx_liab = full_values_liab.index(max_val_liab)
        max_date_liab = full_dates_liab[max_idx_liab]
        min_val_liab = min(full_values_liab)
        min_idx_liab = full_values_liab.index(min_val_liab)
        min_date_liab = full_dates_liab[min_idx_liab]

        col_bar_liab, col_area_liab = st.columns(2)

        with col_bar_liab:
            fig_bar_liab = go.Figure()
            fig_bar_liab.add_trace(go.Bar(
                x=full_dates_liab,
                y=full_values_liab,
                marker_color='rgba(4, 68, 119, 0.7)',
                hovertemplate='%{x|%Y-%m-%d}: %{y:,.0f}<extra></extra>'
            ))
            fig_bar_liab.add_trace(go.Scatter(
                x=[last_date_liab],
                y=[last_val_liab],
                mode='markers+text',
                marker=dict(size=10, color='rgb(216, 12, 24)'),
                text=[f'{last_val_liab:,.0f}'],
                textposition='top center',
                textfont=dict(size=9, color='rgb(216, 12, 24)'),
                showlegend=False
            ))
            fig_bar_liab.add_trace(go.Scatter(
                x=[max_date_liab],
                y=[max_val_liab],
                mode='markers+text',
                marker=dict(size=8, color='rgb(40, 160, 100)'),
                text=[f'最高: {max_val_liab:,.0f}'],
                textposition='top center',
                textfont=dict(size=8, color='rgb(40, 160, 100)'),
                showlegend=False
            ))
            fig_bar_liab.add_trace(go.Scatter(
                x=[min_date_liab],
                y=[min_val_liab],
                mode='markers+text',
                marker=dict(size=8, color='rgb(94, 94, 94)'),
                text=[f'最低: {min_val_liab:,.0f}'],
                textposition='bottom center',
                textfont=dict(size=8, color='rgb(94, 94, 94)'),
                showlegend=False
            ))
            fig_bar_liab.update_layout(
                title=dict(text=f"{trend_liab_display.strip()} - 柱状图（绝对规模）", font=dict(size=12)),
                xaxis=dict(title='日期'),
                yaxis=dict(title='规模 (百万美元)'),
                showlegend=False,
                margin=dict(t=40, b=30, l=50, r=20),
                height=300
            )
            st.plotly_chart(fig_bar_liab, width="stretch")

        with col_area_liab:
            total_liab_values = df_liab_filtered['总负债'].fillna(0).tolist()
            percent_values_liab = [(v / t * 100) if t != 0 else 0 for v, t in zip(full_values_liab, total_liab_values)]
            last_percent_liab = percent_values_liab[-1]

            fig_area_liab = go.Figure()
            fig_area_liab.add_trace(go.Scatter(
                x=full_dates_liab,
                y=percent_values_liab,
                fill='tozeroy',
                mode='lines',
                line=dict(color='rgb(4, 68, 119)', width=1.5),
                fillcolor='rgba(4, 68, 119, 0.2)',
                hovertemplate='%{x|%Y-%m-%d}: %{y:.2f}%<extra></extra>'
            ))
            fig_area_liab.add_trace(go.Scatter(
                x=[last_date_liab],
                y=[last_percent_liab],
                mode='markers+text',
                marker=dict(size=10, color='rgb(216, 12, 24)'),
                text=[f'{last_percent_liab:.2f}%'],
                textposition='top center',
                textfont=dict(size=9, color='rgb(216, 12, 24)'),
                showlegend=False
            ))
            fig_area_liab.update_layout(
                title=dict(text=f"{trend_liab_display.strip()} - 面积图（占总负债%）", font=dict(size=12)),
                xaxis=dict(title='日期'),
                yaxis=dict(title='占比 (%)'),
                showlegend=False,
                margin=dict(t=40, b=30, l=50, r=20),
                height=300
            )
            st.plotly_chart(fig_area_liab, width="stretch")

        st.markdown("---")

        # ========== ②周度变动分析 ==========
        st.markdown("#### ②周度变动分析")

        weekly_liab_options = [(col, display) for col, display, level in liab_items if col in df_liab_filtered.columns]
        weekly_liab_idx = st.selectbox(
            "选择负债项目",
            range(len(weekly_liab_options)),
            format_func=lambda x: weekly_liab_options[x][1],
            key='weekly_change_liab'
        )
        weekly_liab_col = weekly_liab_options[weekly_liab_idx][0]
        weekly_liab_display = weekly_liab_options[weekly_liab_idx][1]

        weekly_changes_liab = []
        weekly_dates_liab = []
        values_liab = df_liab_filtered[weekly_liab_col].fillna(0).tolist()
        all_dates_liab_chart = df_liab_filtered['日期'].tolist()

        for i in range(1, len(values_liab)):
            change = values_liab[i] - values_liab[i-1]
            weekly_changes_liab.append(change)
            weekly_dates_liab.append(all_dates_liab_chart[i])

        fig_bar_liab2 = go.Figure()
        colors_bar_liab = ['rgb(216, 12, 24)' if v < 0 else 'rgb(4, 68, 119)' for v in weekly_changes_liab]
        fig_bar_liab2.add_trace(go.Bar(
            x=weekly_dates_liab,
            y=weekly_changes_liab,
            marker_color=colors_bar_liab,
            hovertemplate='%{x|%Y-%m-%d}: %{y:,.0f}<extra></extra>'
        ))
        last_change_liab = weekly_changes_liab[-1]
        last_date_liab2 = weekly_dates_liab[-1]
        fig_bar_liab2.add_trace(go.Scatter(
            x=[last_date_liab2],
            y=[last_change_liab],
            mode='markers+text',
            marker=dict(size=10, color='rgb(216, 12, 24)' if last_change_liab < 0 else 'rgb(4, 68, 119)'),
            text=[f'{last_change_liab:+,.0f}'],
            textposition='top center' if last_change_liab >= 0 else 'bottom center',
            textfont=dict(size=9, color='rgb(216, 12, 24)' if last_change_liab < 0 else 'rgb(4, 68, 119)'),
            showlegend=False
        ))
        max_change_liab = max(weekly_changes_liab)
        max_idx_liab2 = weekly_changes_liab.index(max_change_liab)
        max_date_liab2 = weekly_dates_liab[max_idx_liab2]
        fig_bar_liab2.add_trace(go.Scatter(
            x=[max_date_liab2],
            y=[max_change_liab],
            mode='markers+text',
            marker=dict(size=8, color='rgb(40, 160, 100)'),
            text=[f'最大增: {max_change_liab:,.0f}'],
            textposition='top center',
            textfont=dict(size=8, color='rgb(40, 160, 100)'),
            showlegend=False
        ))
        min_change_liab = min(weekly_changes_liab)
        min_idx_liab2 = weekly_changes_liab.index(min_change_liab)
        min_date_liab2 = weekly_dates_liab[min_idx_liab2]
        fig_bar_liab2.add_trace(go.Scatter(
            x=[min_date_liab2],
            y=[min_change_liab],
            mode='markers+text',
            marker=dict(size=8, color='rgb(94, 94, 94)'),
            text=[f'最大减: {min_change_liab:,.0f}'],
            textposition='bottom center',
            textfont=dict(size=8, color='rgb(94, 94, 94)'),
            showlegend=False
        ))
        fig_bar_liab2.update_layout(
            title=dict(text=f"{weekly_liab_display.strip()} 周度变动", font=dict(size=13)),
            xaxis=dict(title='日期'),
            yaxis=dict(title='变动量 (百万美元)'),
            showlegend=False,
            margin=dict(t=40, b=30, l=50, r=20),
            height=350
        )
        st.plotly_chart(fig_bar_liab2, width="stretch")

        st.markdown("---")

        # ========== ③环比增速 ==========
        st.markdown("#### ③环比增速")

        growth_liab_options = [(col, display) for col, display, level in liab_items if col in df_liab_filtered.columns]
        growth_liab_idx = st.selectbox(
            "选择负债项目",
            range(len(growth_liab_options)),
            format_func=lambda x: growth_liab_options[x][1],
            key='growth_liab_select'
        )
        growth_liab_col = growth_liab_options[growth_liab_idx][0]
        growth_liab_display = growth_liab_options[growth_liab_idx][1]

        values_liab2 = df_liab_filtered[growth_liab_col].fillna(0).tolist()
        all_dates_liab2 = df_liab_filtered['日期'].tolist()

        growth_rates_liab = []
        growth_dates_liab = []
        for i in range(1, len(values_liab2)):
            if values_liab2[i-1] != 0:
                rate = (values_liab2[i] - values_liab2[i-1]) / abs(values_liab2[i-1]) * 100
            else:
                rate = 0 if values_liab2[i] == 0 else 100
            growth_rates_liab.append(rate)
            growth_dates_liab.append(all_dates_liab2[i])

        fig_growth_liab = go.Figure()
        colors_growth_liab = ['rgb(216, 12, 24)' if v < 0 else 'rgb(4, 68, 119)' for v in growth_rates_liab]
        fig_growth_liab.add_trace(go.Bar(
            x=growth_dates_liab,
            y=growth_rates_liab,
            marker_color=colors_growth_liab,
            hovertemplate='%{x|%Y-%m-%d}: %{y:.2f}%<extra></extra>'
        ))
        last_rate_liab = growth_rates_liab[-1]
        last_date_liab3 = growth_dates_liab[-1]
        fig_growth_liab.add_trace(go.Scatter(
            x=[last_date_liab3],
            y=[last_rate_liab],
            mode='markers+text',
            marker=dict(size=10, color='rgb(216, 12, 24)' if last_rate_liab < 0 else 'rgb(4, 68, 119)'),
            text=[f'{last_rate_liab:+.2f}%'],
            textposition='top center' if last_rate_liab >= 0 else 'bottom center',
            textfont=dict(size=9, color='rgb(216, 12, 24)' if last_rate_liab < 0 else 'rgb(4, 68, 119)'),
            showlegend=False
        ))
        fig_growth_liab.add_hline(y=0, line_dash='dash', line_color='gray', opacity=0.5)
        fig_growth_liab.update_layout(
            title=dict(text=f"{growth_liab_display.strip()} 环比增速", font=dict(size=13)),
            xaxis=dict(title='日期'),
            yaxis=dict(title='环比增速 (%)'),
            showlegend=False,
            margin=dict(t=40, b=30, l=50, r=20),
            height=300
        )
        st.plotly_chart(fig_growth_liab, width="stretch")

    # ==================== 流动性分析 ====================
    with tab3:
        st.markdown('<div class="chart-title">流动性分析</div>', unsafe_allow_html=True)

        # ==================== 蓄水池 ====================
        st.markdown("### 🏊 蓄水池")

        with st.expander("📋 为什么要看蓄水池？", expanded=True):
            st.markdown("""
            把联储资产负债表想象成一个大水库，联储放出来的钱在三个"水池"里流转：

            - **准备金**：银行把多余的钱存回联储账户，赚IORB利率（现在约4.4%）。这是最大的那块。
            - **ON RRP逆回购**：货币市场基金把钱临时停在联储过夜。2022–2023年曾膨胀到近2万亿，现已基本归零。
            - **TGA财政存款**：政府的"支票账户"，收税时钱流进来，花钱时流入银行准备金。

            **为什么要看三者之和？** 单看任何一个会被误导。比如准备金涨了——但可能是TGA在花钱，钱只是从财政账户倒进银行准备金。三者加总才能看出联储有没有真正净投放或收回流动性。
            """)

        # 计算蓄水池数据
        df_pools = df_liab[['日期', '存款-其他存款', '逆向回购协议', '财政部一般账户']].copy()
        df_pools = df_pools.rename(columns={
            '存款-其他存款': '准备金',
            '逆向回购协议': 'ON RRP',
            '财政部一般账户': 'TGA'
        })
        df_pools['总蓄水池'] = df_pools['准备金'] + df_pools['ON RRP'] + df_pools['TGA']

        # 筛选日期范围
        df_pools_filtered = df_pools[(df_pools['日期'] >= start_date) & (df_pools['日期'] <= end_date)].copy()
        df_pools_filtered = df_pools_filtered.sort_values('日期')

        # 最新数据
        latest_pool = df_pools_filtered.iloc[-1]
        pool_date_str = latest_pool['日期'].strftime('%Y-%m-%d')

        # 蓄水池堆叠面积图
        fig_pools = go.Figure()

        # 准备金（底层，红色）
        fig_pools.add_trace(go.Scatter(
            x=df_pools_filtered['日期'],
            y=df_pools_filtered['准备金'],
            mode='lines',
            name='准备金',
            line=dict(color='rgb(216, 12, 24)', width=2),
            stackgroup='one',
            fillcolor='rgba(216, 12, 24, 0.8)',
            hovertemplate='准备金: %{y:,.0f}<extra></extra>'
        ))

        # ON RRP（中间层，灰色）
        fig_pools.add_trace(go.Scatter(
            x=df_pools_filtered['日期'],
            y=df_pools_filtered['ON RRP'],
            mode='lines',
            name='ON RRP',
            line=dict(color='rgb(148, 148, 149)', width=2),
            stackgroup='one',
            fillcolor='rgba(148, 148, 149, 0.5)',
            hovertemplate='ON RRP: %{y:,.0f}<extra></extra>'
        ))

        # TGA（顶层，蓝色）
        fig_pools.add_trace(go.Scatter(
            x=df_pools_filtered['日期'],
            y=df_pools_filtered['TGA'],
            mode='lines',
            name='TGA',
            line=dict(color='rgb(4, 68, 119)', width=2),
            stackgroup='one',
            fillcolor='rgba(4, 68, 119, 0.5)',
            hovertemplate='TGA: %{y:,.0f}<extra></extra>'
        ))

        # 添加最新数值标注
        fig_pools.add_annotation(
            x=latest_pool['日期'],
            y=latest_pool['准备金'] / 2,
            text=f"准备金<br>{latest_pool['准备金']/1000:,.0f}K",
            showarrow=False,
            font=dict(color='white', size=11, family='SimHei'),
            bgcolor='rgba(216, 12, 24, 0.8)',
            bordercolor='rgba(216, 12, 24, 0.8)',
            borderwidth=2,
            borderpad=4
        )

        fig_pools.add_annotation(
            x=latest_pool['日期'],
            y=latest_pool['准备金'] + latest_pool['ON RRP'] / 2,
            text=f"ON RRP<br>{latest_pool['ON RRP']/1000:,.0f}K",
            showarrow=False,
            font=dict(color='white', size=11, family='SimHei'),
            bgcolor='rgba(148, 148, 149, 0.8)',
            bordercolor='rgba(148, 148, 149, 0.8)',
            borderwidth=2,
            borderpad=4
        )

        fig_pools.add_annotation(
            x=latest_pool['日期'],
            y=latest_pool['准备金'] + latest_pool['ON RRP'] + latest_pool['TGA'] / 2,
            text=f"TGA<br>{latest_pool['TGA']/1000:,.0f}K",
            showarrow=False,
            font=dict(color='white', size=11, family='SimHei'),
            bgcolor='rgba(4, 68, 119, 0.8)',
            bordercolor='rgba(4, 68, 119, 0.8)',
            borderwidth=2,
            borderpad=4
        )

        fig_pools.update_layout(
            title=f'联储负债端三大蓄水池 ({pool_date_str}) | 总量: {latest_pool["总蓄水池"]/1000:,.0f}K 百万美元',
            height=400,
            hovermode='x unified',
            plot_bgcolor='#fafafa',
            paper_bgcolor='white',
            font=dict(family='SimHei', size=12),
            margin=dict(t=60, b=30, l=50, r=20),
            legend=dict(orientation='h', y=1.02, x=0.5, xanchor='center')
        )
        fig_pools.update_xaxes(title_text="日期", linecolor='#333', linewidth=0.5, ticks='inside')
        fig_pools.update_yaxes(title_text="金额 (百万美元)", linecolor='#333', linewidth=0.5, ticks='inside')

        st.plotly_chart(fig_pools, width="stretch")

        st.markdown("---")

        # ==================== 净流动性 ====================
        st.markdown("### 💧 净流动性")

        # 指标说明
        with st.expander("📋 指标说明：为什么净流动性能代表市场流动性？", expanded=True):
            st.markdown("""
            **净流动性 = 联储总资产 - 逆回购协议(RRP) - 美国财政存款(TGA)**

            - **总资产**：联储持有的国债、MBS等，代表注入市场的流动性
            - **RRP**：金融机构存放在联储的资金，暂时退出流通，需扣除
            - **TGA**：财政部在联储的存款，进入TGA即抽离银行体系流动性

            净流动性直接反映银行体系可用准备金规模，与SOFR、回购市场利率密切相关，是监测市场流动性压力的关键指标。
            """)

        # 计算净流动性
        df_liquidity = df_asset[['日期', '总资产']].copy()
        df_liquidity = df_liquidity.merge(df_liab[['日期', '逆向回购协议', '财政部一般账户']], on='日期', how='left')
        df_liquidity['净流动性'] = df_liquidity['总资产'] - df_liquidity['逆向回购协议'] - df_liquidity['财政部一般账户']

        # 根据日期范围筛选
        df_liquidity_filtered = df_liquidity[(df_liquidity['日期'] >= start_date) & (df_liquidity['日期'] <= end_date)].copy()
        df_liquidity_filtered = df_liquidity_filtered.sort_values('日期')

        # 日期选择器
        all_dates_liq = df_liquidity['日期'].sort_values(ascending=False).tolist()
        date_options_liq = []
        for d in all_dates_liq:
            if pd.notna(d):
                week_num = d.isocalendar()[1]
                date_str = d.strftime('%Y-%m-%d')
                date_options_liq.append(f"{d.year}年第{week_num}周({date_str})")

        col_date1, col_date2 = st.columns(2)
        with col_date1:
            liq_date_idx = st.selectbox(
                "选择日期",
                range(len(date_options_liq)),
                format_func=lambda x: date_options_liq[x],
                index=0,
                key='liq_date_selector'
            )
        with col_date2:
            base_period_options_liq = ["上一周", "一个月前", "三个月前", "一年前", "五年前", "十年前", "自定义"]
            base_period_sel_liq = st.selectbox(
                "选择基期",
                base_period_options_liq,
                index=0,
                key='liq_base_period'
            )

        # 处理基期选择
        if base_period_sel_liq == "自定义":
            base_date_idx_liq = st.selectbox(
                "选择基期日期",
                range(len(date_options_liq)),
                format_func=lambda x: date_options_liq[x],
                index=1,
                key='liq_base_date_selector'
            )
            base_idx_liq = len(df_liquidity) - 1 - base_date_idx_liq
        else:
            selected_date_liq = all_dates_liq[liq_date_idx]
            selected_idx_liq = df_liquidity[df_liquidity['日期'] == selected_date_liq].index[0] if len(df_liquidity[df_liquidity['日期'] == selected_date_liq]) > 0 else len(df_liquidity) - 1

            if base_period_sel_liq == "上一周":
                base_idx_liq = selected_idx_liq - 1 if selected_idx_liq > 0 else selected_idx_liq
            elif base_period_sel_liq == "一个月前":
                base_idx_liq = max(0, selected_idx_liq - 4)
            elif base_period_sel_liq == "三个月前":
                base_idx_liq = max(0, selected_idx_liq - 13)
            elif base_period_sel_liq == "一年前":
                base_idx_liq = max(0, selected_idx_liq - 52)
            elif base_period_sel_liq == "五年前":
                base_idx_liq = max(0, selected_idx_liq - 260)
            elif base_period_sel_liq == "十年前":
                base_idx_liq = max(0, selected_idx_liq - 520)
            else:
                base_idx_liq = max(0, selected_idx_liq - 1)

        selected_date_liq = all_dates_liq[liq_date_idx]
        selected_liq_row = df_liquidity[df_liquidity['日期'] == selected_date_liq].iloc[0] if len(df_liquidity[df_liquidity['日期'] == selected_date_liq]) > 0 else df_liquidity.iloc[-1]

        selected_idx_liq = df_liquidity[df_liquidity['日期'] == selected_date_liq].index[0] if len(df_liquidity[df_liquidity['日期'] == selected_date_liq]) > 0 else len(df_liquidity) - 1
        if base_period_sel_liq == "自定义":
            base_idx_liq = len(df_liquidity) - 1 - base_date_idx_liq
        prev_liq_row = df_liquidity.iloc[base_idx_liq]

        base_date_display_liq = df_liquidity.iloc[base_idx_liq]['日期'].strftime('%Y-%m-%d')
        st.markdown(f"<div style='color: #666; font-size: 12px; margin-bottom: 10px;'>对比基期: {base_date_display_liq} ({base_period_sel_liq})</div>", unsafe_allow_html=True)

        # 关键指标卡片
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            net_liq = selected_liq_row['净流动性']
            prev_net_liq = prev_liq_row['净流动性']
            change_liq = net_liq - prev_net_liq
            change_pct_liq = change_liq / prev_net_liq * 100 if prev_net_liq != 0 else 0
            st.metric("净流动性 (百万美元)", f"{net_liq:,.0f}", f"{change_liq:+,.0f} ({change_pct_liq:+.2f}%)")

        with col2:
            total_asset_liq = selected_liq_row['总资产']
            prev_total_asset = prev_liq_row['总资产']
            change_asset_liq = total_asset_liq - prev_total_asset
            change_asset_pct = change_asset_liq / prev_total_asset * 100 if prev_total_asset != 0 else 0
            st.metric("联储总资产", f"{total_asset_liq:,.0f}", f"{change_asset_liq:+,.0f} ({change_asset_pct:+.2f}%)")

        with col3:
            rrp = selected_liq_row['逆向回购协议']
            prev_rrp = prev_liq_row['逆向回购协议']
            change_rrp = rrp - prev_rrp
            change_rrp_pct = change_rrp / prev_rrp * 100 if prev_rrp != 0 else 0
            st.metric("逆回购协议(RRP)", f"{rrp:,.0f}", f"{change_rrp:+,.0f} ({change_rrp_pct:+.2f}%)")

        with col4:
            tga = selected_liq_row['财政部一般账户']
            prev_tga = prev_liq_row['财政部一般账户']
            change_tga = tga - prev_tga
            change_tga_pct = change_tga / prev_tga * 100 if prev_tga != 0 else 0
            st.metric("财政存款(TGA)", f"{tga:,.0f}", f"{change_tga:+,.0f} ({change_tga_pct:+.2f}%)")

        st.markdown("---")

        # ========== ①净流动性时间趋势 ==========
        st.markdown("#### ①净流动性时间趋势")

        fig_liq_trend = make_subplots(rows=1, cols=2, subplot_titles=('净流动性规模', '净流动性占总资产比重'))

        # 左图：净流动性规模
        fig_liq_trend.add_trace(go.Scatter(
            x=df_liquidity_filtered['日期'],
            y=df_liquidity_filtered['净流动性'],
            mode='lines',
            name='净流动性',
            line=dict(color='rgb(4, 68, 119)', width=2),
            fill='tozeroy',
            fillcolor='rgba(4, 68, 119, 0.2)',
            hovertemplate='%{y:,.0f}<extra></extra>'
        ), row=1, col=1)

        # 计算净流动性占比
        df_liquidity_filtered['净流动性占比'] = df_liquidity_filtered['净流动性'] / df_liquidity_filtered['总资产'] * 100

        # 右图：净流动性占比
        fig_liq_trend.add_trace(go.Scatter(
            x=df_liquidity_filtered['日期'],
            y=df_liquidity_filtered['净流动性占比'],
            mode='lines',
            name='占比',
            line=dict(color='rgb(216, 12, 24)', width=2),
            fill='tozeroy',
            fillcolor='rgba(216, 12, 24, 0.2)',
            hovertemplate='%{y:.1f}%<extra></extra>'
        ), row=1, col=2)

        fig_liq_trend.update_layout(
            height=350,
            showlegend=False,
            hovermode='x unified',
            plot_bgcolor='#fafafa',
            paper_bgcolor='white',
            font=dict(family='SimHei', size=12),
            margin=dict(t=40, b=30, l=50, r=20)
        )
        fig_liq_trend.update_xaxes(title_text="日期", linecolor='#333', linewidth=0.5, ticks='inside')
        fig_liq_trend.update_yaxes(title_text="金额 (百万美元)", row=1, col=1, linecolor='#333', linewidth=0.5, ticks='inside')
        fig_liq_trend.update_yaxes(title_text="占比 (%)", row=1, col=2, linecolor='#333', linewidth=0.5, ticks='inside')

        st.plotly_chart(fig_liq_trend, width="stretch")

        # ========== ②流动性构成分解 ==========
        st.markdown("#### ②流动性构成分解")

        col_left, col_right = st.columns(2)

        with col_left:
            # 左图：堆叠面积图展示净流动性 = 总资产 - RRP - TGA
            fig_decompose = go.Figure()

            # 净流动性（底层，正数）
            fig_decompose.add_trace(go.Scatter(
                x=df_liquidity_filtered['日期'],
                y=df_liquidity_filtered['净流动性'],
                mode='lines',
                name='净流动性',
                line=dict(color='rgb(4, 68, 119)', width=2),
                stackgroup='pos',
                fillcolor='rgba(4, 68, 119, 0.4)',
                hovertemplate='%{y:,.0f}<extra></extra>'
            ))

            # RRP（中间层）
            fig_decompose.add_trace(go.Scatter(
                x=df_liquidity_filtered['日期'],
                y=df_liquidity_filtered['逆向回购协议'],
                mode='lines',
                name='RRP',
                line=dict(color='rgb(216, 12, 24)', width=2),
                stackgroup='pos',
                fillcolor='rgba(216, 12, 24, 0.4)',
                hovertemplate='%{y:,.0f}<extra></extra>'
            ))

            # TGA（顶层）
            fig_decompose.add_trace(go.Scatter(
                x=df_liquidity_filtered['日期'],
                y=df_liquidity_filtered['财政部一般账户'],
                mode='lines',
                name='TGA',
                line=dict(color='rgb(148, 148, 149)', width=2),
                stackgroup='pos',
                fillcolor='rgba(148, 148, 149, 0.4)',
                hovertemplate='%{y:,.0f}<extra></extra>'
            ))

            fig_decompose.update_layout(
                title='流动性构成时序 (净流动性+RRP+TGA=总资产)',
                height=400,
                hovermode='x unified',
                plot_bgcolor='#fafafa',
                paper_bgcolor='white',
                font=dict(family='SimHei', size=12),
                margin=dict(t=50, b=30, l=50, r=20),
                legend=dict(orientation='h', y=1.05, x=0.5, xanchor='center')
            )
            fig_decompose.update_xaxes(title_text="日期", linecolor='#333', linewidth=0.5, ticks='inside')
            fig_decompose.update_yaxes(title_text="金额 (百万美元)", linecolor='#333', linewidth=0.5, ticks='inside')
            st.plotly_chart(fig_decompose, width="stretch")

        with col_right:
            # 右图：瀑布图展示计算（使用顶部选择的日期）
            fig_waterfall = go.Figure(go.Waterfall(
                name='流动性构成',
                orientation='v',
                measure=['relative', 'relative', 'relative', 'total'],
                x=['总资产', '- RRP', '- TGA', '净流动性'],
                y=[selected_liq_row['总资产'], -selected_liq_row['逆向回购协议'], -selected_liq_row['财政部一般账户'], 0],
                connector={'line': {'color': 'rgb(94, 94, 94)'}},
                increasing={'marker': {'color': 'rgb(4, 68, 119)'}},
                decreasing={'marker': {'color': 'rgb(216, 12, 24)'}},
                totals={'marker': {'color': 'rgb(40, 160, 100)'}},
                text=[f'{selected_liq_row["总资产"]:,.0f}',
                      f'-{selected_liq_row["逆向回购协议"]:,.0f}',
                      f'-{selected_liq_row["财政部一般账户"]:,.0f}',
                      f'{net_liq:,.0f}'],
                textposition='outside'
            ))

            fig_waterfall.update_layout(
                title='流动性计算分解',
                height=400,
                showlegend=False,
                plot_bgcolor='#fafafa',
                paper_bgcolor='white',
                font=dict(family='SimHei', size=11),
                margin=dict(t=50, b=30, l=60, r=20)
            )
            fig_waterfall.update_yaxes(title_text="金额 (百万美元)", linecolor='#333', linewidth=0.5, ticks='inside')
            st.plotly_chart(fig_waterfall, width="stretch")

        # ========== ③周度变动分析 ==========
        st.markdown("#### ③周度变动分析")

        df_liq_weekly = df_liquidity_filtered.copy()
        df_liq_weekly['净流动性变动'] = df_liq_weekly['净流动性'].diff()
        df_liq_weekly['总资产变动'] = df_liq_weekly['总资产'].diff()
        df_liq_weekly['RRP变动'] = df_liq_weekly['逆向回购协议'].diff()
        df_liq_weekly['TGA变动'] = df_liq_weekly['财政部一般账户'].diff()

        fig_weekly_liq = make_subplots(rows=2, cols=1, subplot_titles=('净流动性周度变动', '构成项周度变动'))

        colors_liq_weekly = ['#2ca02c' if x >= 0 else '#d62728' for x in df_liq_weekly['净流动性变动'].fillna(0)]

        fig_weekly_liq.add_trace(go.Bar(
            x=df_liq_weekly['日期'],
            y=df_liq_weekly['净流动性变动'],
            name='净流动性变动',
            marker_color=colors_liq_weekly
        ), row=1, col=1)

        fig_weekly_liq.add_trace(go.Bar(
            x=df_liq_weekly['日期'],
            y=df_liq_weekly['总资产变动'],
            name='总资产变动',
            marker_color='rgb(4, 68, 119)'
        ), row=2, col=1)

        fig_weekly_liq.add_trace(go.Bar(
            x=df_liq_weekly['日期'],
            y=-df_liq_weekly['RRP变动'],
            name='RRP变动(反向)',
            marker_color='rgb(216, 12, 24)'
        ), row=2, col=1)

        fig_weekly_liq.add_trace(go.Bar(
            x=df_liq_weekly['日期'],
            y=-df_liq_weekly['TGA变动'],
            name='TGA变动(反向)',
            marker_color='rgb(148, 148, 149)'
        ), row=2, col=1)

        fig_weekly_liq.update_layout(
            height=500,
            hovermode='x unified',
            plot_bgcolor='#fafafa',
            paper_bgcolor='white',
            font=dict(family='SimHei', size=12),
            margin=dict(t=40, b=30, l=50, r=20),
            barmode='group',
            legend=dict(orientation='h', y=1.02, x=0.5, xanchor='center')
        )
        fig_weekly_liq.update_xaxes(title_text="日期", row=2, col=1, linecolor='#333', linewidth=0.5)
        fig_weekly_liq.update_yaxes(title_text="金额 (百万美元)", row=1, col=1, linecolor='#333', linewidth=0.5)
        fig_weekly_liq.update_yaxes(title_text="金额 (百万美元)", row=2, col=1, linecolor='#333', linewidth=0.5)

        st.plotly_chart(fig_weekly_liq, width="stretch")

        # ========== ④环比增速 ==========
        st.markdown("#### ④环比增速")

        df_liq_weekly['净流动性增速'] = df_liq_weekly['净流动性'].pct_change() * 100

        fig_growth_liq = go.Figure()
        fig_growth_liq.add_trace(go.Scatter(
            x=df_liq_weekly['日期'],
            y=df_liq_weekly['净流动性增速'],
            mode='lines',
            name='净流动性环比增速',
            line=dict(color='rgb(4, 68, 119)', width=2),
            hovertemplate='%{y:.2f}%<extra></extra>'
        ))
        fig_growth_liq.add_hline(y=0, line_dash='dash', line_color='gray', opacity=0.5)

        fig_growth_liq.update_layout(
            height=300,
            hovermode='x unified',
            plot_bgcolor='#fafafa',
            paper_bgcolor='white',
            font=dict(family='SimHei', size=12),
            margin=dict(t=40, b=30, l=50, r=20)
        )
        fig_growth_liq.update_xaxes(title_text="日期", linecolor='#333', linewidth=0.5)
        fig_growth_liq.update_yaxes(title_text="增速 (%)", linecolor='#333', linewidth=0.5)

        st.plotly_chart(fig_growth_liq, width="stretch")

        # ========== ⑤历史对比表 ==========
        st.markdown("#### ⑤历史对比")

        periods_liq = [
            ("一周前", 1),
            ("一个月前", 4),
            ("三个月前", 13),
            ("一年前", 52),
            ("五年前", 260),
            ("十年前", 520)
        ]

        liq_table_data = []
        current_idx = df_liquidity[df_liquidity['日期'] == selected_date_liq].index[0] if len(df_liquidity[df_liquidity['日期'] == selected_date_liq]) > 0 else len(df_liquidity) - 1

        for period_name, weeks in periods_liq:
            past_idx = current_idx - weeks
            if past_idx >= 0 and past_idx < len(df_liquidity):
                past_row = df_liquidity.iloc[past_idx]
                change_val = net_liq - past_row['净流动性']
                change_pct_val = change_val / past_row['净流动性'] * 100 if past_row['净流动性'] != 0 else 0
                liq_table_data.append({
                    '时间段': period_name,
                    '日期': past_row['日期'].strftime('%Y-%m-%d'),
                    '净流动性': f"{past_row['净流动性']:,.0f}",
                    '变动': f"{change_val:+,.0f}",
                    '变动百分比': f"{change_pct_val:+.2f}%"
                })
            else:
                liq_table_data.append({
                    '时间段': period_name,
                    '日期': '-',
                    '净流动性': '-',
                    '变动': '-',
                    '变动百分比': '-'
                })

        df_liq_table = pd.DataFrame(liq_table_data)
        st.dataframe(df_liq_table, use_container_width=True, hide_index=True)

    # ==================== 详细数据 ====================
    with tab4:
        st.markdown('<div class="chart-title">详细数据</div>', unsafe_allow_html=True)

        data_type = st.radio("选择数据类型", ["资产数据", "负债数据"], horizontal=True)

        if data_type == "资产数据":
            st.markdown(f"**资产数据明细** (共 {len(df_asset_filtered)} 条记录)")
            st.dataframe(
                df_asset_filtered.style.format({
                    col: '{:,.0f}' for col in df_asset_filtered.columns if col != '日期'
                }),
                use_container_width=True,
                height=400
            )

            csv = df_asset_filtered.to_csv(index=False).encode('utf-8-sig')
            st.download_button(
                label="📥 下载资产数据 (CSV)",
                data=csv,
                file_name=f"fed_asset_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv",
                mime='text/csv'
            )
        else:
            st.markdown(f"**负债数据明细** (共 {len(df_liab_filtered)} 条记录)")
            st.dataframe(
                df_liab_filtered.style.format({
                    col: '{:,.0f}' for col in df_liab_filtered.columns if col != '日期'
                }),
                use_container_width=True,
                height=400
            )

            csv = df_liab_filtered.to_csv(index=False).encode('utf-8-sig')
            st.download_button(
                label="📥 下载负债数据 (CSV)",
                data=csv,
                file_name=f"fed_liability_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv",
                mime='text/csv'
            )

    # 页脚
    st.markdown("---")
    footer_col1, footer_col2, footer_col3 = st.columns([1, 2, 1])
    with footer_col2:
        st.markdown(f"""
        <div style="text-align: center; padding: 10px;">
            <p style="color: #888; font-size: 12px; margin: 0;">数据来源: 国信宏观 | 更新日期: {datetime.now().strftime('%Y-%m-%d')}</p>
        </div>
        """, unsafe_allow_html=True)
